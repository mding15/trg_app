"""
run_equity_model.py — Run the equity model pipeline.

Reads the Securities tab from data/maintenance/Excel/EquityModel.xlsx, fetches historical prices
from the HDF market data store via utils.mkt_data.get_market_data(), and calls
models.equity_model.run_model().

Date range is taken from the model Parameters.csv (TS Start Date / TS End Date).
If model_id is not supplied, the default is read from utils.var_utils.get_default_model_id().

For securities with no price data in the HDF store, mkt_data_extract.extract_yh_price()
is called to fetch from YH. Securities that still have no data after the fetch are
skipped and logged.

Usage:
    python run_equity_model.py
    python run_equity_model.py --model-id M_20251231
    python run_equity_model.py --model-id M_20251231 --submodel-id Equity.5
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

SCRIPT_DIR    = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR.parent))  # trg_app/ — required for models.equity_model imports

from _paths import EXCEL_DIR
WORKBOOK_PATH = EXCEL_DIR / 'EquityModel.xlsx'


# ── logging ────────────────────────────────────────────────────────────────────

def _setup_logger(task: str) -> logging.Logger:
    log_dir = SCRIPT_DIR.parent.parent / 'log'
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f'run_equity_model_{task}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'

    logger = logging.getLogger(f'run_equity_model_{task}')
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter('%(asctime)s  %(levelname)-8s  %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return logger


# ── helpers ────────────────────────────────────────────────────────────────────

def _read_securities(wb: openpyxl.Workbook) -> list[dict]:
    """Read the Securities tab. Returns list of {SecurityID, Ticker, ...} dicts in tab order."""
    ws = wb['Securities']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if row.get('SecurityID'):
            rows.append(row)
    return rows


# ── task: run-model ────────────────────────────────────────────────────────────

def run_equity_model(model_id: str | None, submodel_id: str | None, logger: logging.Logger) -> None:
    from models import equity_model, model_utils
    from utils import mkt_data, var_utils
    from mkt_data import mkt_data_extract

    # ── Resolve model_id and date range ───────────────────────────────────────
    if model_id is None:
        model_id = var_utils.get_default_model_id()
    model_params = model_utils.read_Model_Parameters(model_id)
    from_date = pd.to_datetime(model_params['TS Start Date'])
    to_date   = pd.to_datetime(model_params['TS End Date'])

    logger.info(
        f"=== run-model started: workbook='{WORKBOOK_PATH}' "
        f"model_id={model_id} submodel_id={submodel_id or '(auto)'} "
        f"date_range={from_date.date()} → {to_date.date()} ==="
    )

    # ── Read Securities tab ────────────────────────────────────────────────────
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    if 'Securities' not in wb.sheetnames:
        raise ValueError("'Securities' tab not found in workbook")

    sec_rows   = _read_securities(wb)
    securities = pd.DataFrame(sec_rows)
    sec_ids    = securities['SecurityID'].tolist()
    logger.info(f"Read {len(securities)} securities from 'Securities' tab")

    # ── Fetch prices from HDF ──────────────────────────────────────────────────
    hist_prices = mkt_data.get_market_data(sec_ids, from_date=from_date, to_date=to_date)

    present     = set(hist_prices.columns) if not hist_prices.empty else set()
    missing_ids = [s for s in sec_ids if s not in present or hist_prices[s].isna().all()]

    if missing_ids:
        logger.warning(f"{len(missing_ids)} security_id(s) with no price data — fetching from YH: {missing_ids}")
        mkt_data_extract.extract_yh_price(security_ids=missing_ids)
        hist_prices = mkt_data.get_market_data(sec_ids, from_date=from_date, to_date=to_date)

        still_missing = [s for s in missing_ids if s not in hist_prices.columns or hist_prices[s].isna().all()]
        if still_missing:
            logger.warning(f"{len(still_missing)} security_id(s) still have no data after fetch — skipping: {still_missing}")
            securities  = securities[~securities['SecurityID'].isin(still_missing)]
            hist_prices = hist_prices[[c for c in hist_prices.columns if c not in still_missing]]

    # zero prices produce -inf log returns — treat as missing
    hist_prices = hist_prices.replace(0, np.nan)

    logger.info(
        f"Price matrix: {len(hist_prices)} dates × {len(hist_prices.columns)} securities "
        f"({hist_prices.index.min().date()} → {hist_prices.index.max().date()})"
    )

    # ── Run model ──────────────────────────────────────────────────────────────
    logger.info("Calling equity_model.run_model() ...")
    DATA = equity_model.run_model(securities, hist_prices, model_id, submodel_id)

    params = DATA.get('Parameters', {})
    logger.info(
        f"run_model() complete: model_id={params.get('Model ID')} "
        f"submodel_id={params.get('Submodel ID')}"
    )
    logger.info("=== run-model done ===")


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Run the equity model pipeline.')
    parser.add_argument('--model-id', default=None, metavar='MODEL_ID',
                        help='Model ID (default: from var_utils.get_default_model_id())')
    parser.add_argument('--submodel-id', default=None, metavar='SUBMODEL_ID',
                        help='Submodel ID (default: auto-generated as Equity.<timestamp>)')
    args = parser.parse_args()

    _logger = _setup_logger('run_model')
    run_equity_model(args.model_id, args.submodel_id, _logger)
