"""
figi_create_security.py — Create new securities via OpenFIGI lookup.

Reads a list of securities from an Excel file, skips any whose identifiers are
already in security_xref, calls OpenFIGI for FIGI data, and for each match:
  - inserts a row in security_info  (SecurityName, AssetClass, AssetType,
                                      Currency derived from ISIN prefix)
    … OR links to an existing security_id when figi/comp_figi/shareclass_figi
       already appears in security_lookup (FIGI-sharing)
  - adds ISIN, CUSIP, BB_GLOBAL, and Ticker entries to security_xref
  - upserts a row in security_lookup (preserves existing isin/cusip/name/etc.
    on conflict; always overwrites comp_figi and shareclass_figi)

Results are written to a 'Results' sheet in the same Excel file.
All DB writes are committed in a single transaction.

Identifier priority (first non-blank wins):
    ISIN → CUSIP → FIGI (BB_GLOBAL) → Ticker

Skip logic (symmetric — any hit causes the row to be skipped):
    - ISIN already in security_xref as REF_TYPE='ISIN'
    - CUSIP already in security_xref as REF_TYPE='CUSIP'
    - FIGI already in security_xref as REF_TYPE='BB_GLOBAL'  (when no ISIN/CUSIP)
    - Ticker already in security_xref as REF_TYPE='Ticker'   (when no ISIN/CUSIP/FIGI)

Input sheet columns:
    security_name   (required)
    isin            (optional — preferred identifier; looked up as ID_ISIN)
    cusip           (optional — second choice; looked up as ID_CUSIP)
    figi            (optional — composite FIGI / BB_GLOBAL; looked up as ID_BB_GLOBAL)
    ticker          (optional — last fallback; looked up as TICKER)
    exchange        (optional — exchange code for ticker lookup, e.g. 'US', 'LN')

Public API:
    create(rows, dry_run, log) -> list[dict]   # callable from other modules
    run(file, sheet, dry_run)                  # CLI entry point

Usage:
    python maintenance/figi_create_security.py
    python maintenance/figi_create_security.py --file figi_create_security.xlsx --sheet Securities
    python maintenance/figi_create_security.py --dry-run
"""
from __future__ import annotations

import argparse
import csv
import logging
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'process2'))

from database2 import pg_connection
from figi_utils import FIGI_COLUMNS, fetch, pick_representative
from security_utils import add_xref_if_missing, create_security

EXCEL_DIR     = Path(__file__).resolve().parent / 'Excel'
DEFAULT_FILE  = 'figi_create_security.xlsx'
DEFAULT_SHEET = 'Securities'
DATA_SOURCE   = 'FIGI'

_ISIN_CURRENCY: dict[str, str] = {
    'US': 'USD', 'CA': 'CAD', 'GB': 'GBP', 'IE': 'EUR',
    'CH': 'CHF', 'LU': 'EUR', 'NL': 'EUR', 'DE': 'EUR',
    'FR': 'EUR', 'JP': 'JPY', 'HK': 'HKD', 'AU': 'AUD',
    'JE': 'GBP', 'GG': 'GBP', 'IM': 'GBP',
    'BM': 'USD', 'KY': 'USD', 'VG': 'USD',
}

_UPSERT_SQL = """
    INSERT INTO security_lookup
        (security_id, name, ticker, exch, isin, cusip, sedol,
         figi, comp_figi, shareclass_figi,
         sectype, sectype2, mkt_sector, update_at)
    VALUES
        (%(security_id)s, %(name)s, %(ticker)s, %(exch)s, %(isin)s, %(cusip)s, %(sedol)s,
         %(figi)s, %(comp_figi)s, %(shareclass_figi)s,
         %(sectype)s, %(sectype2)s, %(mkt_sector)s, NOW())
    ON CONFLICT (figi) DO UPDATE SET
        security_id     = COALESCE(NULLIF(EXCLUDED.security_id,     ''), security_lookup.security_id),
        name            = COALESCE(NULLIF(EXCLUDED.name,            ''), security_lookup.name),
        ticker          = COALESCE(NULLIF(EXCLUDED.ticker,          ''), security_lookup.ticker),
        exch            = COALESCE(NULLIF(EXCLUDED.exch,            ''), security_lookup.exch),
        isin            = COALESCE(NULLIF(EXCLUDED.isin,            ''), security_lookup.isin),
        cusip           = COALESCE(NULLIF(EXCLUDED.cusip,           ''), security_lookup.cusip),
        sedol           = COALESCE(NULLIF(EXCLUDED.sedol,           ''), security_lookup.sedol),
        comp_figi       = EXCLUDED.comp_figi,
        shareclass_figi = EXCLUDED.shareclass_figi,
        sectype         = COALESCE(NULLIF(EXCLUDED.sectype,         ''), security_lookup.sectype),
        sectype2        = COALESCE(NULLIF(EXCLUDED.sectype2,        ''), security_lookup.sectype2),
        mkt_sector      = COALESCE(NULLIF(EXCLUDED.mkt_sector,      ''), security_lookup.mkt_sector),
        update_at       = NOW()
"""


# ── Helpers ───────────────────────────────────────────────────────────────────

_EQUITY_SECTYPE_MAP: dict[str, str] = {
    'ADR':            'Stock',
    'Closed-End Fund':'Fund',
    'Common Stock':   'Stock',
    'ETP':            'ETP',
    'Fund of Funds':  'Fund',
    'NY Reg Shrs':    'Stock',
    'Open-End Fund':  'Fund',
    'REIT':           'REIT',
}


def _map_asset(market_sector: str, security_type: str) -> tuple[str | None, str | None]:
    """Map OpenFIGI marketSector + securityType to (asset_class, asset_type)."""
    if market_sector == 'Corp':
        return 'Bond', 'Bond'
    if market_sector == 'Equity':
        return 'Equity', _EQUITY_SECTYPE_MAP.get(security_type, security_type) or None
    if market_sector == 'Govt':
        asset_type = 'Treasury' if security_type == 'US GOVERNMENT' else 'Sovereign'
        return 'Bond', asset_type
    return market_sector or None, security_type or None


def _setup_logger() -> logging.Logger:
    logger = logging.getLogger('figi_create_security')
    logger.setLevel(logging.DEBUG)
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(
        logging.Formatter('%(asctime)s  %(levelname)-8s  %(message)s', '%H:%M:%S')
    )
    logger.addHandler(handler)
    return logger


def _clean(val) -> str:
    try:
        if pd.isna(val):
            return ''
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def _trunc(value, max_len: int) -> str:
    s = value if isinstance(value, str) else (str(value) if value is not None and str(value) != 'nan' else '')
    return s[:max_len]


def _cusip_to_isin(cusip: str, country_code: str = 'US') -> str:
    base = country_code + cusip.upper()
    digits = ''
    for ch in base:
        if ch.isdigit():
            digits += ch
        elif ch.isalpha():
            digits += str(ord(ch) - ord('A') + 10)
    digits += '0'
    total = 0
    for i, d in enumerate(reversed(digits)):
        n = int(d)
        if i % 2 == 1:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    return base + str((10 - (total % 10)) % 10)


def _load_excel(filepath: Path, sheet: str, log: logging.Logger) -> list[dict]:
    log.info(f"Reading sheet '{sheet}' from {filepath.name}")
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
    except ValueError as e:
        log.error(f"Could not read sheet '{sheet}': {e}")
        sys.exit(1)
    except Exception as e:
        log.error(f"Failed to open Excel file: {e}")
        sys.exit(1)

    if df.empty:
        log.warning('Sheet is empty — nothing to process.')
        sys.exit(0)

    log.info(f"  {len(df)} rows · {len(df.columns)} columns")

    if 'security_name' not in df.columns:
        log.error("Required column 'security_name' not found")
        sys.exit(1)

    for col in ('isin', 'cusip', 'figi', 'ticker', 'exchange'):
        if col not in df.columns:
            df[col] = ''

    return [{col: _clean(row[col]) for col in df.columns} for _, row in df.iterrows()]


def _batch_check_existing(cur, rows: list[dict]) -> set[tuple]:
    """Return set of (REF_TYPE, REF_ID) pairs already present in security_xref."""
    isins   = [r['isin']   for r in rows if r.get('isin')]
    cusips  = [r['cusip']  for r in rows if r.get('cusip')]
    figis   = [r['figi']   for r in rows if r.get('figi')]
    tickers = [r['ticker'] for r in rows if r.get('ticker')]
    existing: set[tuple] = set()

    for ref_type, vals in (
        ('ISIN',     isins),
        ('CUSIP',    cusips),
        ('BB_GLOBAL', figis),
        ('Ticker',   tickers),
    ):
        if vals:
            cur.execute(
                'SELECT "REF_TYPE", "REF_ID" FROM security_xref WHERE "REF_TYPE" = %s AND "REF_ID" = ANY(%s)',
                (ref_type, vals),
            )
            existing.update(cur.fetchall())

    return existing


def _to_db_dict(security_id: str, figi_row: pd.Series, isin: str, cusip: str) -> dict:
    return {
        'security_id':     _trunc(security_id, 20),
        'name':            _trunc(figi_row.get('name'), 255),
        'ticker':          _trunc(figi_row.get('ticker'), 50),
        'exch':            _trunc(figi_row.get('exchCode'), 10),
        'isin':            _trunc(isin, 12),
        'cusip':           _trunc(cusip, 9),
        'sedol':           '',
        'figi':            _trunc(figi_row.get('figi'), 12),
        'comp_figi':       _trunc(figi_row.get('compositeFIGI'), 12),
        'shareclass_figi': _trunc(figi_row.get('shareClassFIGI'), 12),
        'sectype':         _trunc(figi_row.get('securityType'), 100),
        'sectype2':        _trunc(figi_row.get('securityType2'), 100),
        'mkt_sector':      _trunc(figi_row.get('marketSector'), 50),
    }


def _fetch_figi_security_ids(
    cur,
    figi_vals: list[str],
    comp_vals: list[str],
    shareclass_vals: list[str],
) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    """
    Batch-query security_lookup for existing security_ids matching FIGI values.
    Returns (figi_map, comp_map, shareclass_map), each mapping a value to a security_id.
    """
    def _query(col: str, vals: list[str]) -> dict[str, str]:
        vals = [v for v in vals if v]
        if not vals:
            return {}
        cur.execute(
            f'SELECT {col}, security_id FROM security_lookup WHERE {col} = ANY(%s)',
            (vals,),
        )
        return {row[0]: str(row[1]) for row in cur.fetchall()}

    return (
        _query('figi',            figi_vals),
        _query('comp_figi',       comp_vals),
        _query('shareclass_figi', shareclass_vals),
    )


def _resolve_security_id(
    figi_row: pd.Series,
    figi_map: dict[str, str],
    comp_map: dict[str, str],
    shareclass_map: dict[str, str],
    log: logging.Logger,
    name: str = '',
) -> str | None:
    """
    Check figi → comp_figi → shareclass_figi (priority order) against security_lookup.
    Returns the matching security_id, or None if no match found.
    Logs a warning when different FIGI IDs disagree on security_id.
    """
    figi_val       = _trunc(figi_row.get('figi'),           12)
    comp_val       = _trunc(figi_row.get('compositeFIGI'),  12)
    shareclass_val = _trunc(figi_row.get('shareClassFIGI'), 12)

    matches: dict[str, str] = {}
    if figi_val       and figi_val       in figi_map:        matches['figi']            = figi_map[figi_val]
    if comp_val       and comp_val       in comp_map:        matches['comp_figi']       = comp_map[comp_val]
    if shareclass_val and shareclass_val in shareclass_map:  matches['shareclass_figi'] = shareclass_map[shareclass_val]

    if not matches:
        return None

    if len(set(matches.values())) > 1:
        log.warning(f"  FIGI ID conflict for '{name}': {matches} — using figi priority")

    return matches.get('figi') or matches.get('comp_figi') or matches.get('shareclass_figi')


def _write_results(xlsx_path: Path, result_rows: list[dict], log: logging.Logger) -> None:
    headers = ['status', 'security_id', 'security_name', 'figi_name', 'isin', 'cusip',
               'asset_class', 'asset_type', 'currency', 'figi', 'compositeFIGI']
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        if 'Results' in wb.sheetnames:
            del wb['Results']
        ws = wb.create_sheet('Results')
        ws.append(headers)
        for row in result_rows:
            ws.append([row.get(h) for h in headers])
        wb.save(xlsx_path)
        log.info(f"Results written to 'Results' sheet in {xlsx_path.name}")
    except PermissionError:
        csv_path = xlsx_path.with_suffix('.csv')
        log.warning(f"Excel file is locked — writing results to {csv_path.name} instead")
        with csv_path.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(result_rows)
        log.info(f"Results written to {csv_path.name}")


# ── Core logic ────────────────────────────────────────────────────────────────

def create(
    rows: list[dict],
    dry_run: bool = False,
    log: logging.Logger | None = None,
) -> list[dict]:
    """
    Look up ISINs/CUSIPs via OpenFIGI and create new securities in the database.

    Skips rows whose ISIN or CUSIP already exists in security_xref.
    For each match, inserts security_info, security_xref (ISIN, CUSIP, Ticker),
    and upserts security_lookup. All writes are committed in one transaction.

    Each input dict must have:
        security_name  (str, required) — used as SecurityName; falls back to OpenFIGI name if blank
        isin           (str, optional) — preferred identifier; looked up as ID_ISIN
        cusip          (str, optional) — second choice; looked up as ID_CUSIP when isin is blank
        figi           (str, optional) — composite FIGI (BB_GLOBAL); used when isin and cusip are blank; looked up as ID_BB_GLOBAL
        ticker         (str, optional) — last fallback when isin, cusip, and figi are all blank; looked up as TICKER
        exchange       (str, optional) — exchange code for ticker lookup (e.g. 'US', 'LN'); ignored for isin/cusip/figi rows

    Values must already be clean strings (stripped, no NaN). Missing keys default to ''.

    Example input:
        [
            {'security_name': 'Apple Inc',            'isin': 'US0378331005', 'cusip': '037833100', 'figi': '',             'ticker': '',     'exchange': ''},
            {'security_name': 'Tesla Inc',             'isin': 'US88160R1014', 'cusip': '88160R101', 'figi': '',             'ticker': '',     'exchange': ''},
            {'security_name': 'Vanguard Total Bond',   'isin': '',             'cusip': '922908769', 'figi': '',             'ticker': '',     'exchange': ''},
            {'security_name': 'iShares MSCI EM',       'isin': '',             'cusip': '',           'figi': 'BBG000BCZS13', 'ticker': '',     'exchange': ''},
            {'security_name': 'Nvidia Corp',           'isin': '',             'cusip': '',           'figi': '',             'ticker': 'NVDA', 'exchange': 'US'},
            {'security_name': 'BP PLC',                'isin': '',             'cusip': '',           'figi': '',             'ticker': 'BP',   'exchange': 'LN'},
        ]

    Returns one result dict per input row. Possible status values:
        'created'                    — new security created; security_id is populated
        'linked (existing security)' — figi/comp_figi/shareclass_figi matched an existing
                                       security_lookup row; identifiers added to xref under
                                       that security_id, no new security_info row created
        'would create'               — dry_run=True, would have created a new security
        'would link (existing security)' — dry_run=True, would have linked to existing security
        'skipped (already in xref)'  — ISIN, CUSIP, or Ticker already in security_xref
        'skipped (no identifier)'    — isin, cusip, and ticker are all blank
        'skipped (no OpenFIGI data)' — OpenFIGI returned no match

    Example return:
        [
            {'status': 'created', 'security_id': 'T10000042', 'security_name': 'Apple Inc',
             'figi_name': 'APPLE INC', 'isin': 'US0378331005', 'cusip': '037833100',
             'asset_class': 'Equity', 'asset_type': 'Common Stock', 'currency': 'USD',
             'figi': 'BBG000B9XRY4', 'compositeFIGI': 'BBG000B9Y5X2'},
            {'status': 'linked (existing security)', 'security_id': 'T10000007',
             'security_name': 'Apple Inc ADR', 'isin': 'US0378331006', 'cusip': '037833101',
             'figi': 'BBG000B9XRY5', 'compositeFIGI': 'BBG000B9Y5X2'},
            {'status': 'skipped (already in xref)', 'security_name': 'Tesla Inc',
             'isin': 'US88160R1014', 'cusip': '88160R101'},
        ]
    """
    if log is None:
        log = _setup_logger()

    if dry_run:
        log.info('─' * 60)
        log.info('DRY RUN — no data will be written to the database')

    # ── Batch skip check ──────────────────────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            existing = _batch_check_existing(cur, rows)

    log.info(f"{len(existing)} identifier(s) already in security_xref")

    # ── Filter rows and build OpenFIGI id_pairs ───────────────────────────────
    isin_pairs:    list[tuple[str, str]] = []
    cusip_pairs:   list[tuple[str, str]] = []
    figi_pairs:    list[tuple[str, str]] = []
    ticker_pairs:  list[tuple[str, str]] = []
    ticker_extras: list[dict]            = []
    to_process:    list[dict]            = []
    result_rows:   list[dict]            = []

    for i, r in enumerate(rows, 1):
        isin     = r.get('isin',          '')
        cusip    = r.get('cusip',         '')
        figi     = r.get('figi',          '')
        ticker   = r.get('ticker',        '')
        exchange = r.get('exchange',      '')
        name     = r.get('security_name', '')

        if not isin and not cusip and not figi and not ticker:
            log.warning(f"Row {i}: SKIP — no ISIN, CUSIP, FIGI, or Ticker  ({name})")
            result_rows.append({'status': 'skipped (no identifier)', 'security_name': name})
            continue

        if isin and ('ISIN', isin) in existing:
            log.warning(f"Row {i}: SKIP — ISIN '{isin}' already in xref  ({name})")
            result_rows.append({'status': 'skipped (already in xref)', 'security_name': name, 'isin': isin, 'cusip': cusip})
            continue

        if cusip and ('CUSIP', cusip) in existing:
            log.warning(f"Row {i}: SKIP — CUSIP '{cusip}' already in xref  ({name})")
            result_rows.append({'status': 'skipped (already in xref)', 'security_name': name, 'isin': isin, 'cusip': cusip})
            continue

        if not isin and not cusip and figi and ('BB_GLOBAL', figi) in existing:
            log.warning(f"Row {i}: SKIP — FIGI '{figi}' already in xref  ({name})")
            result_rows.append({'status': 'skipped (already in xref)', 'security_name': name, 'figi': figi})
            continue

        if not isin and not cusip and not figi and ticker and ('Ticker', ticker) in existing:
            log.warning(f"Row {i}: SKIP — Ticker '{ticker}' already in xref  ({name})")
            result_rows.append({'status': 'skipped (already in xref)', 'security_name': name, 'ticker': ticker})
            continue

        to_process.append(r)
        if isin:
            isin_pairs.append((isin, isin))
        elif cusip:
            cusip_pairs.append((cusip, _cusip_to_isin(cusip)))
        elif figi:
            figi_pairs.append((figi, figi))
        else:
            ticker_pairs.append((ticker, ticker))
            ticker_extras.append({'exchCode': exchange.upper()} if exchange else {})

    if not to_process:
        log.info('Nothing to process after skip check.')
        return result_rows

    log.info(f"{len(to_process)} row(s) to process")

    # ── Fetch OpenFIGI ────────────────────────────────────────────────────────
    frames: list[pd.DataFrame] = []
    with requests.Session() as session:
        if isin_pairs:
            log.info(f"Fetching {len(isin_pairs)} ISIN(s) from OpenFIGI …")
            frames.append(fetch(isin_pairs, 'ID_ISIN', session, log))
        if cusip_pairs:
            log.info(f"Fetching {len(cusip_pairs)} CUSIP(s) from OpenFIGI …")
            frames.append(fetch(cusip_pairs, 'ID_CUSIP', session, log))
        if figi_pairs:
            log.info(f"Fetching {len(figi_pairs)} FIGI(s) from OpenFIGI …")
            frames.append(fetch(figi_pairs, 'ID_BB_GLOBAL', session, log))
        if ticker_pairs:
            log.info(f"Fetching {len(ticker_pairs)} Ticker(s) from OpenFIGI …")
            frames.append(fetch(ticker_pairs, 'TICKER', session, log, job_extras=ticker_extras))

    long_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=FIGI_COLUMNS)
    rep_df  = pick_representative(long_df) if not long_df.empty else pd.DataFrame(columns=FIGI_COLUMNS)

    figi_by_isin: dict[str, pd.Series] = {}
    for _, frow in rep_df.iterrows():
        key = str(frow.get('isin', ''))
        if key:
            figi_by_isin[key] = frow

    log.info(f"{len(figi_by_isin)} representative FIGI row(s) returned")

    # ── Create securities and upsert security_lookup ──────────────────────────
    created = 0
    linked  = 0
    no_figi = 0

    with pg_connection() as conn:
        with conn.cursor() as cur:
            # Batch-check which FIGI values already exist in security_lookup
            figi_vals       = [_trunc(fr.get('figi'),          12) for fr in figi_by_isin.values()]
            comp_vals       = [_trunc(fr.get('compositeFIGI'), 12) for fr in figi_by_isin.values()]
            shareclass_vals = [_trunc(fr.get('shareClassFIGI'),12) for fr in figi_by_isin.values()]
            figi_map, comp_map, shareclass_map = _fetch_figi_security_ids(
                cur, figi_vals, comp_vals, shareclass_vals,
            )
            log.info(
                f"{len(figi_map)} figi / {len(comp_map)} comp_figi / "
                f"{len(shareclass_map)} shareclass_figi match(es) in security_lookup"
            )

            for r in to_process:
                isin  = r.get('isin',  '')
                cusip = r.get('cusip', '')
                name  = r.get('security_name', '')

                input_figi = r.get('figi', '')
                isin_key   = isin if isin else (_cusip_to_isin(cusip) if cusip else (input_figi if input_figi else r.get('ticker', '')))
                figi_row = figi_by_isin.get(isin_key)

                if figi_row is None:
                    log.warning(f"  SKIP — no OpenFIGI data  ({name}  isin='{isin}'  cusip='{cusip}')")
                    result_rows.append({
                        'status': 'skipped (no OpenFIGI data)',
                        'security_name': name, 'isin': isin, 'cusip': cusip,
                    })
                    no_figi += 1
                    continue

                security_name = name or _trunc(figi_row.get('name'), 1000)
                asset_class, asset_type = _map_asset(
                    _trunc(figi_row.get('marketSector'), 20),
                    _trunc(figi_row.get('securityType'), 20),
                )
                currency      = _ISIN_CURRENCY.get(isin_key[:2].upper(), '')

                existing_id = _resolve_security_id(figi_row, figi_map, comp_map, shareclass_map, log, name)

                if dry_run:
                    if existing_id:
                        log.info(
                            f"  WOULD LINK {existing_id} — '{security_name}'"
                            f"  isin='{isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                        )
                        status = 'would link (existing security)'
                        linked += 1
                    else:
                        log.info(
                            f"  WOULD CREATE — '{security_name}'"
                            f"  currency='{currency}' asset_class='{asset_class}' asset_type='{asset_type}'"
                            f"  isin='{isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                        )
                        status = 'would create'
                        created += 1
                    result_rows.append({
                        'status':        status,
                        'security_id':   existing_id or '',
                        'security_name': security_name,
                        'figi_name':     figi_row.get('name'),
                        'isin':          isin,
                        'cusip':         cusip,
                        'asset_class':   asset_class,
                        'asset_type':    asset_type,
                        'currency':      currency,
                        'figi':          figi_row.get('figi'),
                        'compositeFIGI': figi_row.get('compositeFIGI'),
                    })
                    continue

                if existing_id:
                    security_id = existing_id
                    status      = 'linked (existing security)'
                    linked += 1
                    log.info(
                        f"  LINKED {security_id} — '{security_name}'"
                        f"  isin='{isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                    )
                else:
                    security_id = create_security(cur, security_name, currency, asset_class, asset_type, DATA_SOURCE)
                    status      = 'created'
                    created += 1
                    log.info(
                        f"  CREATED {security_id} — '{security_name}'"
                        f"  isin='{isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                    )

                # security_xref (both created and linked)
                ticker    = _trunc(figi_row.get('ticker'),   50)
                exch_code = _trunc(figi_row.get('exchCode'), 10)
                add_xref_if_missing(cur, security_id, 'ISIN',      isin,       DATA_SOURCE, exch_code)
                add_xref_if_missing(cur, security_id, 'CUSIP',     cusip,      DATA_SOURCE, exch_code)
                add_xref_if_missing(cur, security_id, 'BB_GLOBAL', input_figi, DATA_SOURCE, exch_code)
                add_xref_if_missing(cur, security_id, 'Ticker',    ticker,     DATA_SOURCE, exch_code)

                # security_lookup (both created and linked)
                cur.execute(_UPSERT_SQL, _to_db_dict(security_id, figi_row, isin, cusip))

                result_rows.append({
                    'status':        status,
                    'security_id':   security_id,
                    'security_name': security_name,
                    'figi_name':     figi_row.get('name'),
                    'isin':          isin,
                    'cusip':         cusip,
                    'asset_class':   asset_class,
                    'asset_type':    asset_type,
                    'currency':      currency,
                    'figi':          figi_row.get('figi'),
                    'compositeFIGI': figi_row.get('compositeFIGI'),
                })

        if not dry_run:
            conn.commit()

    log.info('─' * 60)
    skipped = len(rows) - len(to_process)
    log.info(
        f"Done.  Created: {created}  Linked: {linked}  "
        f"No FIGI data: {no_figi}  Skipped: {skipped}  Total: {len(rows)}"
    )
    return result_rows


def run(file: str, sheet: str, dry_run: bool) -> None:
    log        = _setup_logger()
    excel_path = EXCEL_DIR / file

    if not excel_path.exists():
        log.error(f"File not found: {excel_path}")
        sys.exit(1)

    rows        = _load_excel(excel_path, sheet, log)
    result_rows = create(rows, dry_run, log)
    _write_results(excel_path, result_rows, log)


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Create securities via OpenFIGI lookup.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            'Examples:\n'
            '  python maintenance/figi_create_security.py\n'
            '  python maintenance/figi_create_security.py --dry-run\n'
            '  python maintenance/figi_create_security.py --file my_file.xlsx --sheet Sheet1\n'
        ),
    )
    parser.add_argument(
        '--file', default=DEFAULT_FILE, metavar='FILENAME',
        help=f'Excel filename inside maintenance/Excel/ (default: {DEFAULT_FILE})',
    )
    parser.add_argument(
        '--sheet', default=DEFAULT_SHEET, metavar='SHEET',
        help=f'Sheet name to read (default: {DEFAULT_SHEET})',
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Preview what would be created without writing to the database',
    )
    args = parser.parse_args()
    run(args.file, args.sheet, args.dry_run)
