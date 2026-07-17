"""
create_security.py — Create new securities, resolving identifiers via OpenFIGI
when possible and falling back to manually-supplied values otherwise.

For each row in the sheet:
  1. Skips rows with no security_name, or no isin/cusip/figi/ticker at all.
  2. Skips rows whose identifiers already exist in security_xref but point to
     *different* SecurityIDs (self-contradictory input data) — flagged distinctly
     from a normal duplicate, with details in the Results sheet's 'notes' column.
  3. If exactly one existing SecurityID is matched (no conflict), the row is NOT
     skipped — it's treated as a backfill against that existing security, so any
     identifiers the row provides that security_xref doesn't have yet (e.g. sheet
     has ISIN+CUSIP but xref only has CUSIP) still get added, instead of being
     silently lost. See "backfilled" statuses below. If figi_lookup already has
     data for that security_id (from a prior run), it's reused directly instead
     of calling OpenFIGI again — figi_lookup is a local cache of OpenFIGI results.
  4. For backfill rows without a figi_lookup cache hit, and for all new rows,
     tries OpenFIGI using priority order ISIN -> CUSIP -> FIGI -> Ticker.
       - OpenFIGI returns data: derives currency/asset_class/asset_type/name from
         the response (sheet values override when provided). If the row already
         matched an existing security (step 3), backfills that security instead
         of creating/linking a new one; otherwise checks figi_lookup for an
         existing figi/comp_figi/shareclass_figi match and links to that security
         instead of creating a duplicate. Upserts figi_lookup in all three cases.
         DataSource defaults to 'FIGI'.
       - OpenFIGI returns no data: falls back to using whatever currency/
         asset_class/asset_type were supplied in the sheet — currency may be
         blank, in which case it's stored as NULL rather than skipping the row.
         Creates a new security, or backfills an existing one if step 3 matched.
         DataSource defaults to 'MANUAL'.
  5. In all creation/linking/backfilling cases, every non-blank identifier (isin,
     cusip, figi as BB_GLOBAL, ticker) is added to security_xref as provided —
     even a figi value that didn't resolve via OpenFIGI is still recorded.

Results are written to a 'Results' sheet in the same Excel file (falls back to
a CSV if the file is locked). All DB writes for a run are committed in one
transaction.

Sheet columns:
    security_name   (required)
    currency        (optional — manual override / fallback when OpenFIGI has no match)
    asset_class     (optional — manual override / fallback)
    asset_type      (optional — manual override / fallback)
    isin            (optional)
    cusip           (optional)
    figi            (optional — composite FIGI / BB_GLOBAL)
    ticker          (optional)
    exchange        (optional — exchange code, used only when ticker is the lookup identifier)

Public API:
    create(rows, dry_run, log, data_source=None) -> list[dict]   # callable from other modules
    run(file, sheet, dry_run, data_source=None)                  # CLI entry point

Usage:
    python maintenance/create_security.py
    python maintenance/create_security.py --file my_securities.xlsx --sheet Securities
    python maintenance/create_security.py --data-source MANUAL
    python maintenance/create_security.py --dry-run
"""
from __future__ import annotations

import argparse
import csv
import logging
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'process2'))

from database2 import pg_connection
from figi_utils import FIGI_COLUMNS, fetch, pick_representative, FIGI_LOOKUP_UPSERT_SQL as _UPSERT_SQL
from security_utils import add_xref_if_missing, create_security
from _paths import EXCEL_DIR

DEFAULT_FILE  = 'create_security.xlsx'
DEFAULT_SHEET = 'securities'
OPTIONAL_COLS = ['currency', 'asset_class', 'asset_type', 'isin', 'cusip', 'figi', 'ticker', 'exchange']

_ISIN_CURRENCY: dict[str, str] = {
    'US': 'USD', 'CA': 'CAD', 'GB': 'GBP', 'IE': 'EUR',
    'CH': 'CHF', 'LU': 'EUR', 'NL': 'EUR', 'DE': 'EUR',
    'FR': 'EUR', 'JP': 'JPY', 'HK': 'HKD', 'AU': 'AUD',
    'JE': 'GBP', 'GG': 'GBP', 'IM': 'GBP',
    'BM': 'USD', 'KY': 'USD', 'VG': 'USD',
}


# ── Helpers ───────────────────────────────────────────────────────────────────

_EQUITY_SECTYPE_MAP: dict[str, str] = {
    'ADR':            'Stock',
    'Closed-End Fund':'Fund',
    'Common Stock':   'Stock',
    'ETP':            'ETP',
    'Fund of Funds':  'Fund',
    'NY Reg Shrs':    'Stock',
    'Open-End Fund':  'Fund',
    'REIT':           'REIT',
}


def _map_asset(market_sector: str, security_type: str) -> tuple[str | None, str | None]:
    """Map OpenFIGI marketSector + securityType to (asset_class, asset_type)."""
    if market_sector == 'Corp':
        return 'Bond', 'Bond'
    if market_sector == 'Equity':
        return 'Equity', _EQUITY_SECTYPE_MAP.get(security_type, security_type) or None
    if market_sector == 'Govt':
        asset_type = 'Treasury' if security_type == 'US GOVERNMENT' else 'Sovereign'
        return 'Bond', asset_type
    return market_sector or None, security_type or None


def _setup_logger() -> logging.Logger:
    logger = logging.getLogger('create_security')
    logger.setLevel(logging.DEBUG)
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(
        logging.Formatter('%(asctime)s  %(levelname)-8s  %(message)s', '%H:%M:%S')
    )
    logger.addHandler(handler)
    return logger


def _clean(val) -> str:
    try:
        if pd.isna(val):
            return ''
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def _trunc(value, max_len: int) -> str:
    s = value if isinstance(value, str) else (str(value) if value is not None and str(value) != 'nan' else '')
    return s[:max_len]


def _cusip_to_isin(cusip: str, country_code: str = 'US') -> str:
    base = country_code + cusip.upper()
    digits = ''
    for ch in base:
        if ch.isdigit():
            digits += ch
        elif ch.isalpha():
            digits += str(ord(ch) - ord('A') + 10)
    digits += '0'
    total = 0
    for i, d in enumerate(reversed(digits)):
        n = int(d)
        if i % 2 == 1:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    return base + str((10 - (total % 10)) % 10)


def _load_excel(filepath: Path, sheet: str, log: logging.Logger) -> list[dict]:
    log.info(f"Reading sheet '{sheet}' from {filepath.name}")
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
    except ValueError as e:
        log.error(f"Could not read sheet '{sheet}': {e}")
        sys.exit(1)
    except Exception as e:
        log.error(f"Failed to open Excel file: {e}")
        sys.exit(1)

    if df.empty:
        log.warning('Sheet is empty — nothing to process.')
        sys.exit(0)

    log.info(f"  {len(df)} rows · {len(df.columns)} columns")

    if 'security_name' not in df.columns:
        log.error("Required column 'security_name' not found")
        sys.exit(1)

    for col in OPTIONAL_COLS:
        if col not in df.columns:
            df[col] = ''

    if 'cusip' in df.columns:
        prefix = 'CUSIP:'
        is_prefixed = df['cusip'].apply(lambda v: isinstance(v, str) and v.startswith(prefix))
        n_prefixed = int(is_prefixed.sum())
        if n_prefixed:
            df.loc[is_prefixed, 'cusip'] = df.loc[is_prefixed, 'cusip'].str[len(prefix):].str.strip()
            log.info(f"  Stripped '{prefix}' prefix from {n_prefixed} cusip value(s)")

    return [{col: _clean(row[col]) for col in df.columns} for _, row in df.iterrows()]


def _batch_check_existing(cur, rows: list[dict]) -> dict[tuple[str, str], str]:
    """Return {(REF_TYPE, REF_ID): SecurityID} for every identifier already present
    in security_xref. Mapping to SecurityID (rather than just tracking existence)
    lets the caller detect when a row's identifiers point to different securities."""
    isins   = [r['isin']   for r in rows if r.get('isin')]
    cusips  = [r['cusip']  for r in rows if r.get('cusip')]
    figis   = [r['figi']   for r in rows if r.get('figi')]
    tickers = [r['ticker'] for r in rows if r.get('ticker')]
    existing: dict[tuple[str, str], str] = {}

    for ref_type, vals in (
        ('ISIN',      isins),
        ('CUSIP',     cusips),
        ('BB_GLOBAL', figis),
        ('Ticker',    tickers),
    ):
        if vals:
            cur.execute(
                'SELECT "REF_TYPE", "REF_ID", "SecurityID" FROM security_xref WHERE "REF_TYPE" = %s AND "REF_ID" = ANY(%s)',
                (ref_type, vals),
            )
            for rt, rid, sec_id in cur.fetchall():
                existing[(rt, rid)] = str(sec_id)

    return existing


def _to_db_dict(security_id: str, figi_row: pd.Series, isin: str, cusip: str) -> dict:
    return {
        'security_id':     _trunc(security_id, 20),
        'name':            _trunc(figi_row.get('name'), 255),
        'ticker':          _trunc(figi_row.get('ticker'), 50),
        'exch':            _trunc(figi_row.get('exchCode'), 10),
        'isin':            _trunc(isin, 12),
        'cusip':           _trunc(cusip, 9),
        'sedol':           '',
        'figi':            _trunc(figi_row.get('figi'), 12),
        'comp_figi':       _trunc(figi_row.get('compositeFIGI'), 12),
        'shareclass_figi': _trunc(figi_row.get('shareClassFIGI'), 12),
        'sectype':         _trunc(figi_row.get('securityType'), 100),
        'sectype2':        _trunc(figi_row.get('securityType2'), 100),
        'mkt_sector':      _trunc(figi_row.get('marketSector'), 50),
    }


def _fetch_figi_lookup_by_security_id(cur, security_ids: list[str]) -> dict[str, dict]:
    """
    Batch-query figi_lookup for existing rows matching the given security_ids.
    Used to serve backfill rows from cache instead of calling OpenFIGI again for
    a security that's already been resolved before (figi_lookup is a local cache
    of prior OpenFIGI results). Returns {security_id: row_dict}.
    """
    if not security_ids:
        return {}
    cur.execute(
        'SELECT security_id, name, ticker, exch, isin, cusip, sedol, figi, '
        'comp_figi, shareclass_figi, sectype, sectype2, mkt_sector '
        'FROM figi_lookup WHERE security_id = ANY(%s)',
        (security_ids,),
    )
    cols = ['security_id', 'name', 'ticker', 'exch', 'isin', 'cusip', 'sedol', 'figi',
            'comp_figi', 'shareclass_figi', 'sectype', 'sectype2', 'mkt_sector']
    return {str(row[0]): dict(zip(cols, row)) for row in cur.fetchall()}


def _fetch_figi_security_ids(
    cur,
    figi_vals: list[str],
    comp_vals: list[str],
    shareclass_vals: list[str],
) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    """
    Batch-query figi_lookup for existing security_ids matching FIGI values.
    Returns (figi_map, comp_map, shareclass_map), each mapping a value to a security_id.
    """
    def _query(col: str, vals: list[str]) -> dict[str, str]:
        vals = [v for v in vals if v]
        if not vals:
            return {}
        cur.execute(
            f'SELECT {col}, security_id FROM figi_lookup WHERE {col} = ANY(%s)',
            (vals,),
        )
        return {row[0]: str(row[1]) for row in cur.fetchall()}

    return (
        _query('figi',            figi_vals),
        _query('comp_figi',       comp_vals),
        _query('shareclass_figi', shareclass_vals),
    )


def _resolve_security_id(
    figi_row: pd.Series,
    figi_map: dict[str, str],
    comp_map: dict[str, str],
    shareclass_map: dict[str, str],
    log: logging.Logger,
    name: str = '',
) -> str | None:
    """
    Check figi → comp_figi → shareclass_figi (priority order) against figi_lookup.
    Returns the matching security_id, or None if no match found.
    Logs a warning when different FIGI IDs disagree on security_id.
    """
    figi_val       = _trunc(figi_row.get('figi'),           12)
    comp_val       = _trunc(figi_row.get('compositeFIGI'),  12)
    shareclass_val = _trunc(figi_row.get('shareClassFIGI'), 12)

    matches: dict[str, str] = {}
    if figi_val       and figi_val       in figi_map:        matches['figi']            = figi_map[figi_val]
    if comp_val       and comp_val       in comp_map:        matches['comp_figi']       = comp_map[comp_val]
    if shareclass_val and shareclass_val in shareclass_map:  matches['shareclass_figi'] = shareclass_map[shareclass_val]

    if not matches:
        return None

    if len(set(matches.values())) > 1:
        log.warning(f"  FIGI ID conflict for '{name}': {matches} — using figi priority")

    return matches.get('figi') or matches.get('comp_figi') or matches.get('shareclass_figi')


def _write_results(xlsx_path: Path, result_rows: list[dict], log: logging.Logger) -> None:
    headers = ['status', 'security_id', 'security_name', 'figi_name', 'isin', 'cusip',
               'asset_class', 'asset_type', 'currency', 'figi', 'compositeFIGI', 'notes']
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        if 'Results' in wb.sheetnames:
            del wb['Results']
        ws = wb.create_sheet('Results')
        ws.append(headers)
        for row in result_rows:
            ws.append([row.get(h) for h in headers])
        wb.save(xlsx_path)
        log.info(f"Results written to 'Results' sheet in {xlsx_path.name}")
    except PermissionError:
        csv_path = xlsx_path.with_suffix('.csv')
        log.warning(f"Excel file is locked — writing results to {csv_path.name} instead")
        with csv_path.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(result_rows)
        log.info(f"Results written to {csv_path.name}")


# ── Core logic ────────────────────────────────────────────────────────────────

def create(
    rows: list[dict],
    dry_run: bool = False,
    log: logging.Logger | None = None,
    data_source: str | None = None,
) -> list[dict]:
    """
    Create, link, or backfill securities, trying OpenFIGI first and falling back
    to manually-supplied sheet values when OpenFIGI has no match.

    Skips rows with no security_name, no isin/cusip/figi/ticker at all, or whose
    identifiers already exist in security_xref but point to different SecurityIDs.
    A row whose identifiers all agree with an existing SecurityID is NOT skipped —
    it backfills any identifiers that security is still missing.

    Each input dict must have:
        security_name  (str, required) — used as SecurityName; falls back to OpenFIGI name if blank
        currency       (str, optional) — manual value; used as override (FIGI path) or
                                          as-is / NULL if blank (manual fallback path)
        asset_class    (str, optional) — manual override / fallback
        asset_type     (str, optional) — manual override / fallback
        isin           (str, optional) — preferred identifier; looked up as ID_ISIN
        cusip          (str, optional) — second choice; looked up as ID_CUSIP when isin is blank
        figi           (str, optional) — composite FIGI (BB_GLOBAL); used when isin and cusip are blank; looked up as ID_BB_GLOBAL
        ticker         (str, optional) — last fallback when isin, cusip, and figi are all blank; looked up as TICKER
        exchange       (str, optional) — exchange code for ticker lookup (e.g. 'US', 'LN'); ignored for isin/cusip/figi rows

    Values must already be clean strings (stripped, no NaN). Missing keys default to ''.

    data_source: if given, overrides the DataSource written for every row. If None
    (default), each row gets 'FIGI' when created/linked via an OpenFIGI match, or
    'MANUAL' when created via the manual fallback path.

    Returns one result dict per input row. Possible status values:
        'created (figi)'                  — new security created from OpenFIGI data
        'linked (existing security)'      — figi/comp_figi/shareclass_figi matched an existing
                                             figi_lookup row; identifiers added to xref under
                                             that security_id, no new security_info row created
        'created (manual)'                — OpenFIGI had no match; created from sheet values
                                             (currency may be NULL if left blank)
        'backfilled (figi)'               — row's identifiers already matched an existing
                                             security_xref entry; OpenFIGI resolved data for it,
                                             so any missing identifiers and figi_lookup data are
                                             added to that existing security (no new security_info
                                             row, no figi_lookup dedup check — the direct
                                             security_xref match is authoritative)
        'backfilled (manual)'             — same as above, but OpenFIGI had no match; only the
                                             sheet-provided identifiers are added (no figi_lookup upsert)
        'backfilled (cached)'             — row's identifiers already matched an existing
                                             security_xref entry, AND figi_lookup already has data
                                             for that security_id from a prior run — reused directly,
                                             no OpenFIGI call and no figi_lookup upsert
        'would create (figi)'             — dry_run=True, would have created via OpenFIGI data
        'would link (existing security)'  — dry_run=True, would have linked to existing security
        'would create (manual)'           — dry_run=True, would have created via manual fallback
        'would backfill (figi)'           — dry_run=True, would have backfilled via OpenFIGI data
        'would backfill (manual)'         — dry_run=True, would have backfilled via sheet values
        'would backfill (cached)'         — dry_run=True, would have backfilled via cached figi_lookup data
        'skipped (conflicting identifiers)' — two or more of this row's identifiers already
                                             exist in security_xref but point to *different*
                                             SecurityIDs — the input data is self-contradictory;
                                             see the 'notes' column in the Results sheet for detail
        'skipped (no identifier)'         — isin, cusip, figi, and ticker are all blank
        'skipped (no security_name)'      — security_name is blank
    """
    if log is None:
        log = _setup_logger()

    if dry_run:
        log.info('─' * 60)
        log.info('DRY RUN — no data will be written to the database')

    # ── Batch skip check ──────────────────────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            existing = _batch_check_existing(cur, rows)

    log.info(f"{len(existing)} identifier(s) already in security_xref")

    # ── Pass 1: skip-checks, conflict check, backfill-target detection ────────
    to_process:    list[dict]       = []
    existing_ids:  list[str | None] = []  # parallel to to_process
    result_rows:   list[dict]       = []

    for i, r in enumerate(rows, 1):
        isin   = r.get('isin',          '')
        cusip  = r.get('cusip',         '')
        figi   = r.get('figi',          '')
        ticker = r.get('ticker',        '')
        name   = r.get('security_name', '')

        if not name:
            log.warning(f"Row {i}: SKIP — no security_name")
            result_rows.append({'status': 'skipped (no security_name)'})
            continue

        if not isin and not cusip and not figi and not ticker:
            log.warning(f"Row {i}: SKIP — no ISIN, CUSIP, FIGI, or Ticker  ({name})")
            result_rows.append({'status': 'skipped (no identifier)', 'security_name': name})
            continue

        # Conflict check: if this row's identifiers each already exist in security_xref
        # but point to *different* SecurityIDs, the input data is self-contradictory —
        # flag it explicitly rather than silently skipping under the generic
        # "already in xref" reason below, which would hide the disagreement.
        matched_ids: dict[str, str] = {}
        for ref_type, ref_id in (('ISIN', isin), ('CUSIP', cusip), ('BB_GLOBAL', figi), ('Ticker', ticker)):
            if ref_id and (ref_type, ref_id) in existing:
                matched_ids[ref_type] = existing[(ref_type, ref_id)]

        if len(set(matched_ids.values())) > 1:
            conflict_detail = ', '.join(f"{rt}→{sid}" for rt, sid in matched_ids.items())
            log.warning(f"Row {i}: SKIP — conflicting identifiers point to different securities  ({name})  [{conflict_detail}]")
            result_rows.append({
                'status': 'skipped (conflicting identifiers)',
                'security_name': name, 'isin': isin, 'cusip': cusip, 'figi': figi,
                'notes': conflict_detail,
            })
            continue

        # If exactly one distinct SecurityID matched above (no conflict), this row's
        # security already exists — but the row may still carry identifiers that
        # security_xref doesn't have yet (e.g. sheet has ISIN+CUSIP, xref only has
        # CUSIP). Rather than skipping and losing that data, the row is carried
        # forward so those missing identifiers get backfilled against the existing
        # security below (from cached figi_lookup data if available, else OpenFIGI).
        existing_security_id = next(iter(set(matched_ids.values())), None)

        to_process.append(r)
        existing_ids.append(existing_security_id)

    if not to_process:
        log.info('Nothing to process after skip check.')
        return result_rows

    log.info(f"{len(to_process)} row(s) to process")

    # ── Cache check: figi_lookup is a local cache of prior OpenFIGI results.
    # For backfill rows (existing_security_id already known), reuse cached data
    # instead of calling OpenFIGI again if this security has been resolved before.
    backfill_security_ids = sorted({sid for sid in existing_ids if sid})
    figi_lookup_cache: dict[str, dict] = {}
    if backfill_security_ids:
        with pg_connection() as conn:
            with conn.cursor() as cur:
                figi_lookup_cache = _fetch_figi_lookup_by_security_id(cur, backfill_security_ids)
        log.info(
            f"{len(figi_lookup_cache)} of {len(backfill_security_ids)} backfill security_id(s) "
            f"already cached in figi_lookup — skipping OpenFIGI for those"
        )

    # ── Pass 2: build OpenFIGI id_pairs, skipping rows already served by cache ──
    isin_pairs:    list[tuple[str, str]] = []
    cusip_pairs:   list[tuple[str, str]] = []
    figi_pairs:    list[tuple[str, str]] = []
    ticker_pairs:  list[tuple[str, str]] = []
    ticker_extras: list[dict]            = []

    for r, existing_security_id in zip(to_process, existing_ids):
        if existing_security_id and existing_security_id in figi_lookup_cache:
            continue  # cached backfill — no OpenFIGI call needed

        isin     = r.get('isin',     '')
        cusip    = r.get('cusip',    '')
        figi     = r.get('figi',     '')
        ticker   = r.get('ticker',   '')
        exchange = r.get('exchange', '')
        if isin:
            isin_pairs.append((isin, isin))
        elif cusip:
            cusip_pairs.append((cusip, _cusip_to_isin(cusip)))
        elif figi:
            figi_pairs.append((figi, figi))
        else:
            ticker_pairs.append((ticker, ticker))
            ticker_extras.append({'exchCode': exchange.upper()} if exchange else {})

    # ── Fetch OpenFIGI ────────────────────────────────────────────────────────
    frames: list[pd.DataFrame] = []
    with requests.Session() as session:
        if isin_pairs:
            log.info(f"Fetching {len(isin_pairs)} ISIN(s) from OpenFIGI …")
            frames.append(fetch(isin_pairs, 'ID_ISIN', session, log))
        if cusip_pairs:
            log.info(f"Fetching {len(cusip_pairs)} CUSIP(s) from OpenFIGI …")
            frames.append(fetch(cusip_pairs, 'ID_CUSIP', session, log))
        if figi_pairs:
            log.info(f"Fetching {len(figi_pairs)} FIGI(s) from OpenFIGI …")
            frames.append(fetch(figi_pairs, 'ID_BB_GLOBAL', session, log))
        if ticker_pairs:
            log.info(f"Fetching {len(ticker_pairs)} Ticker(s) from OpenFIGI …")
            frames.append(fetch(ticker_pairs, 'TICKER', session, log, job_extras=ticker_extras))

    long_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=FIGI_COLUMNS)
    rep_df  = pick_representative(long_df) if not long_df.empty else pd.DataFrame(columns=FIGI_COLUMNS)

    figi_by_isin: dict[str, pd.Series] = {}
    for _, frow in rep_df.iterrows():
        key = str(frow.get('isin', ''))
        if key:
            figi_by_isin[key] = frow

    log.info(f"{len(figi_by_isin)} representative FIGI row(s) returned")

    # ── Create/link/backfill securities ──────────────────────────────────────
    created_figi      = 0
    created_manual    = 0
    linked            = 0
    backfilled_figi   = 0
    backfilled_manual = 0
    backfilled_cached = 0

    with pg_connection() as conn:
        with conn.cursor() as cur:
            # Batch-check which FIGI values already exist in figi_lookup
            figi_vals       = [_trunc(fr.get('figi'),          12) for fr in figi_by_isin.values()]
            comp_vals       = [_trunc(fr.get('compositeFIGI'), 12) for fr in figi_by_isin.values()]
            shareclass_vals = [_trunc(fr.get('shareClassFIGI'),12) for fr in figi_by_isin.values()]
            figi_map, comp_map, shareclass_map = _fetch_figi_security_ids(
                cur, figi_vals, comp_vals, shareclass_vals,
            )
            log.info(
                f"{len(figi_map)} figi / {len(comp_map)} comp_figi / "
                f"{len(shareclass_map)} shareclass_figi match(es) in figi_lookup"
            )

            for r, existing_security_id in zip(to_process, existing_ids):
                isin         = r.get('isin',        '')
                cusip        = r.get('cusip',       '')
                input_figi   = r.get('figi',        '')
                input_ticker = r.get('ticker',      '')
                name         = r.get('security_name', '')
                currency_in  = r.get('currency',    '') or None
                asset_class_in = r.get('asset_class', '') or None
                asset_type_in  = r.get('asset_type',  '') or None

                # OpenFIGI's response has no 'isin' field of its own — for a CUSIP-resolved
                # row, the only ISIN available is the one we derive ourselves. For US
                # securities this is the real ISIN (ISO 6166: "US" + CUSIP + check digit),
                # not a guess, so it's safe to record in security_xref/figi_lookup.
                resolved_isin = isin or (_cusip_to_isin(cusip) if cusip else '')

                if existing_security_id and existing_security_id in figi_lookup_cache:
                    # ── Cached backfill path — figi_lookup already has data for this
                    # security from a prior run; reuse it instead of calling OpenFIGI. ──
                    cached        = figi_lookup_cache[existing_security_id]
                    security_id   = existing_security_id
                    cached_ticker = _trunc(cached.get('ticker'), 50)
                    exch_code     = _trunc(cached.get('exch'),   10)
                    row_data_source = data_source or 'FIGI'

                    if dry_run:
                        log.info(
                            f"  WOULD BACKFILL (cached) {security_id} — '{name}'"
                            f"  isin='{resolved_isin}' cusip='{cusip}' figi='{input_figi}'"
                        )
                        backfilled_cached += 1
                        result_rows.append({
                            'status':        'would backfill (cached)',
                            'security_id':   security_id,
                            'security_name': name,
                            'figi_name':     cached.get('name'),
                            'isin':          resolved_isin,
                            'cusip':         cusip,
                            'currency':      currency_in,
                            'figi':          cached.get('figi'),
                            'compositeFIGI': cached.get('comp_figi'),
                        })
                        continue

                    backfilled_cached += 1
                    log.info(
                        f"  BACKFILLED (cached) {security_id} — '{name}'"
                        f"  isin='{resolved_isin}' cusip='{cusip}' figi='{input_figi}'"
                    )

                    add_xref_if_missing(cur, security_id, 'ISIN',      resolved_isin, row_data_source, exch_code)
                    add_xref_if_missing(cur, security_id, 'CUSIP',     cusip,         row_data_source, exch_code)
                    add_xref_if_missing(cur, security_id, 'BB_GLOBAL', input_figi,    row_data_source, exch_code)
                    add_xref_if_missing(cur, security_id, 'Ticker',    cached_ticker, row_data_source, exch_code)
                    # figi_lookup already has this security's data — no upsert (no refresh for now)

                    result_rows.append({
                        'status':        'backfilled (cached)',
                        'security_id':   security_id,
                        'security_name': name,
                        'figi_name':     cached.get('name'),
                        'isin':          resolved_isin,
                        'cusip':         cusip,
                        'currency':      currency_in,
                        'figi':          cached.get('figi'),
                        'compositeFIGI': cached.get('comp_figi'),
                    })
                    continue

                isin_key = isin if isin else (_cusip_to_isin(cusip) if cusip else (input_figi if input_figi else input_ticker))
                figi_row = figi_by_isin.get(isin_key)

                if figi_row is not None:
                    # ── FIGI path ─────────────────────────────────────────────
                    security_name = name or _trunc(figi_row.get('name'), 1000)
                    asset_class_derived, asset_type_derived = _map_asset(
                        _trunc(figi_row.get('marketSector'), 20),
                        _trunc(figi_row.get('securityType'), 20),
                    )
                    currency    = currency_in    or _ISIN_CURRENCY.get(isin_key[:2].upper(), '') or None
                    asset_class = asset_class_in or asset_class_derived
                    asset_type  = asset_type_in  or asset_type_derived
                    row_data_source = data_source or 'FIGI'

                    # If this row's identifiers already matched an existing security in
                    # security_xref (existing_security_id), that's authoritative — no need
                    # to also check figi_lookup for a dedup match.
                    dedup_id = None if existing_security_id else _resolve_security_id(
                        figi_row, figi_map, comp_map, shareclass_map, log, name,
                    )

                    if dry_run:
                        if existing_security_id:
                            log.info(
                                f"  WOULD BACKFILL {existing_security_id} — '{security_name}'"
                                f"  isin='{resolved_isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                            )
                            status = 'would backfill (figi)'
                            backfilled_figi += 1
                        elif dedup_id:
                            log.info(
                                f"  WOULD LINK {dedup_id} — '{security_name}'"
                                f"  isin='{resolved_isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                            )
                            status = 'would link (existing security)'
                            linked += 1
                        else:
                            log.info(
                                f"  WOULD CREATE (figi) — '{security_name}'"
                                f"  currency='{currency}' asset_class='{asset_class}' asset_type='{asset_type}'"
                                f"  isin='{resolved_isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                            )
                            status = 'would create (figi)'
                            created_figi += 1
                        result_rows.append({
                            'status':        status,
                            'security_id':   existing_security_id or dedup_id or '',
                            'security_name': security_name,
                            'figi_name':     figi_row.get('name'),
                            'isin':          resolved_isin,
                            'cusip':         cusip,
                            'asset_class':   asset_class,
                            'asset_type':    asset_type,
                            'currency':      currency,
                            'figi':          figi_row.get('figi'),
                            'compositeFIGI': figi_row.get('compositeFIGI'),
                        })
                        continue

                    if existing_security_id:
                        security_id = existing_security_id
                        status      = 'backfilled (figi)'
                        backfilled_figi += 1
                        log.info(
                            f"  BACKFILLED {security_id} — '{security_name}'"
                            f"  isin='{resolved_isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                        )
                    elif dedup_id:
                        security_id = dedup_id
                        status      = 'linked (existing security)'
                        linked += 1
                        log.info(
                            f"  LINKED {security_id} — '{security_name}'"
                            f"  isin='{resolved_isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                        )
                    else:
                        security_id = create_security(cur, security_name, currency, asset_class, asset_type, row_data_source)
                        status      = 'created (figi)'
                        created_figi += 1
                        log.info(
                            f"  CREATED {security_id} — '{security_name}'"
                            f"  isin='{resolved_isin}' cusip='{cusip}' figi='{figi_row.get('figi')}'"
                        )

                    # security_xref (create / link / backfill) — OpenFIGI's canonical ticker/exchange
                    ticker    = _trunc(figi_row.get('ticker'),   50)
                    exch_code = _trunc(figi_row.get('exchCode'), 10)
                    add_xref_if_missing(cur, security_id, 'ISIN',      resolved_isin, row_data_source, exch_code)
                    add_xref_if_missing(cur, security_id, 'CUSIP',     cusip,         row_data_source, exch_code)
                    add_xref_if_missing(cur, security_id, 'BB_GLOBAL', input_figi,    row_data_source, exch_code)
                    add_xref_if_missing(cur, security_id, 'Ticker',    ticker,        row_data_source, exch_code)

                    # figi_lookup (create / link / backfill)
                    cur.execute(_UPSERT_SQL, _to_db_dict(security_id, figi_row, resolved_isin, cusip))

                    result_rows.append({
                        'status':        status,
                        'security_id':   security_id,
                        'security_name': security_name,
                        'figi_name':     figi_row.get('name'),
                        'isin':          resolved_isin,
                        'cusip':         cusip,
                        'asset_class':   asset_class,
                        'asset_type':    asset_type,
                        'currency':      currency,
                        'figi':          figi_row.get('figi'),
                        'compositeFIGI': figi_row.get('compositeFIGI'),
                    })

                else:
                    # ── Manual fallback path — no OpenFIGI match ─────────────
                    security_name   = name
                    currency        = currency_in
                    asset_class     = asset_class_in
                    asset_type      = asset_type_in
                    row_data_source = data_source or 'MANUAL'

                    if dry_run:
                        if existing_security_id:
                            log.info(
                                f"  WOULD BACKFILL (manual) {existing_security_id} — '{security_name}'"
                                f" isin='{resolved_isin}' cusip='{cusip}' ticker='{input_ticker}'"
                            )
                            status = 'would backfill (manual)'
                            backfilled_manual += 1
                        else:
                            log.info(
                                f"  WOULD CREATE (manual) — '{security_name}' currency='{currency}'"
                                f" asset_class='{asset_class}' isin='{resolved_isin}' cusip='{cusip}' ticker='{input_ticker}'"
                            )
                            status = 'would create (manual)'
                            created_manual += 1
                        result_rows.append({
                            'status':        status,
                            'security_id':   existing_security_id or '',
                            'security_name': security_name,
                            'isin':          resolved_isin,
                            'cusip':         cusip,
                            'asset_class':   asset_class,
                            'asset_type':    asset_type,
                            'currency':      currency,
                            'figi':          input_figi,
                        })
                        continue

                    if existing_security_id:
                        security_id = existing_security_id
                        status      = 'backfilled (manual)'
                        backfilled_manual += 1
                        log.info(f"  BACKFILLED {security_id} (manual) — '{security_name}' isin='{resolved_isin}' cusip='{cusip}'")
                    else:
                        security_id = create_security(cur, security_name, currency, asset_class, asset_type, row_data_source)
                        status      = 'created (manual)'
                        created_manual += 1
                        log.info(f"  CREATED {security_id} (manual) — '{security_name}' isin='{resolved_isin}' cusip='{cusip}'")

                    add_xref_if_missing(cur, security_id, 'ISIN',      resolved_isin, row_data_source)
                    add_xref_if_missing(cur, security_id, 'CUSIP',     cusip,         row_data_source)
                    add_xref_if_missing(cur, security_id, 'BB_GLOBAL', input_figi,    row_data_source)
                    add_xref_if_missing(cur, security_id, 'Ticker',    input_ticker,  row_data_source)

                    result_rows.append({
                        'status':        status,
                        'security_id':   security_id,
                        'security_name': security_name,
                        'isin':          resolved_isin,
                        'cusip':         cusip,
                        'asset_class':   asset_class,
                        'asset_type':    asset_type,
                        'currency':      currency,
                        'figi':          input_figi,
                    })

        if not dry_run:
            conn.commit()

    log.info('─' * 60)
    skipped = len(rows) - len(to_process)
    log.info(
        f"Done.  Created (figi): {created_figi}  Created (manual): {created_manual}  Linked: {linked}  "
        f"Backfilled (figi): {backfilled_figi}  Backfilled (manual): {backfilled_manual}  "
        f"Backfilled (cached): {backfilled_cached}  Skipped: {skipped}  Total: {len(rows)}"
    )
    return result_rows


def run(file: str, sheet: str, dry_run: bool, data_source: str | None = None) -> None:
    log        = _setup_logger()
    excel_path = EXCEL_DIR / file

    if not excel_path.exists():
        log.error(f"File not found: {excel_path}")
        sys.exit(1)

    rows        = _load_excel(excel_path, sheet, log)
    result_rows = create(rows, dry_run, log, data_source)
    _write_results(excel_path, result_rows, log)


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Create or link securities, via OpenFIGI when possible, manual fallback otherwise.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            'Examples:\n'
            '  python maintenance/create_security.py\n'
            '  python maintenance/create_security.py --dry-run\n'
            '  python maintenance/create_security.py --file my_file.xlsx --sheet Sheet1\n'
            '  python maintenance/create_security.py --data-source MANUAL\n'
        ),
    )
    parser.add_argument(
        '--file', default=DEFAULT_FILE, metavar='FILENAME',
        help=f'Excel filename inside data/maintenance/Excel/ (default: {DEFAULT_FILE})',
    )
    parser.add_argument(
        '--sheet', default=DEFAULT_SHEET, metavar='SHEET',
        help=f'Sheet name to read (default: {DEFAULT_SHEET})',
    )
    parser.add_argument(
        '--data-source', default=None, metavar='SOURCE',
        help="DataSource written to security_info and security_xref. If omitted, each row "
             "defaults to 'FIGI' (created/linked via OpenFIGI) or 'MANUAL' (manual fallback).",
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Preview what would be created without writing to the database',
    )
    args = parser.parse_args()
    run(args.file, args.sheet, args.dry_run, args.data_source)
