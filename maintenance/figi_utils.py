"""
Shared OpenFIGI utilities used by security_lookup.py and isin_lookup.py.

Exports:
    OPENFIGI_API_KEY        — API key for OpenFIGI
    OPENFIGI_URL            — OpenFIGI /v3/mapping endpoint
    OPENFIGI_BATCH          — max identifiers per batch request
    HOME_EXCH               — ISIN country prefix -> preferred exchange codes
    FIGI_COLUMNS            — canonical column order for OpenFIGI result DataFrames
    FIGI_LOOKUP_UPSERT_SQL  — shared upsert statement for the figi_lookup table
    fetch                   — call OpenFIGI /v3/mapping for a list of identifiers
    pick_representative     — reduce a long FIGI DataFrame to one row per ISIN
    test                    — read ISINs/CUSIPs from Excel, fetch FIGIs, write Results sheet
"""
from __future__ import annotations

import logging
import sys
import time
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from _paths import EXCEL_DIR

OPENFIGI_API_KEY = '9a5b92a9-cae1-47ad-bb52-9d6293d18364'
OPENFIGI_URL     = "https://api.openfigi.com/v3/mapping"
OPENFIGI_BATCH   = 100

# ISIN country prefix -> preferred home composite exchange codes, in priority order.
HOME_EXCH: dict[str, list[str]] = {
    "US": ["US"],
    "CA": ["CN"],
    "GB": ["LN"],
    "IE": ["ID"],
    "CH": ["SW"],
    "LU": ["LX"],
    "NL": ["NA"],
    "DE": ["GR"],
    "FR": ["FP"],
    "JP": ["JP"],
    "HK": ["HK"],
    "AU": ["AU"],
    # Offshore domiciles -> where such issuers typically list
    "JE": ["LN"],
    "GG": ["LN"],
    "IM": ["LN"],
    "BM": ["US", "HK", "LN"],
    "KY": ["HK", "US", "LN"],
    "VG": ["US", "LN"],
}

FIGI_COLUMNS = [
    "isin", "figi", "name", "ticker", "exchCode", "compositeFIGI",
    "securityType", "marketSector", "shareClassFIGI", "securityType2",
    "securityDescription",
]

# Shared by create_security.py and security_reconcile.py — both upsert into
# figi_lookup keyed on figi, preferring the new value but falling back to the
# existing one when the new value is blank. comp_figi/shareclass_figi/update_at
# always take the new value (FIGI matches are treated as authoritative for those).
FIGI_LOOKUP_UPSERT_SQL = """
    INSERT INTO figi_lookup
        (security_id, name, ticker, exch, isin, cusip, sedol,
         figi, comp_figi, shareclass_figi,
         sectype, sectype2, mkt_sector, update_at)
    VALUES
        (%(security_id)s, %(name)s, %(ticker)s, %(exch)s, %(isin)s, %(cusip)s, %(sedol)s,
         %(figi)s, %(comp_figi)s, %(shareclass_figi)s,
         %(sectype)s, %(sectype2)s, %(mkt_sector)s, NOW())
    ON CONFLICT (figi) DO UPDATE SET
        security_id     = COALESCE(NULLIF(EXCLUDED.security_id,     ''), figi_lookup.security_id),
        name            = COALESCE(NULLIF(EXCLUDED.name,            ''), figi_lookup.name),
        ticker          = COALESCE(NULLIF(EXCLUDED.ticker,          ''), figi_lookup.ticker),
        exch            = COALESCE(NULLIF(EXCLUDED.exch,            ''), figi_lookup.exch),
        isin            = COALESCE(NULLIF(EXCLUDED.isin,            ''), figi_lookup.isin),
        cusip           = COALESCE(NULLIF(EXCLUDED.cusip,           ''), figi_lookup.cusip),
        sedol           = COALESCE(NULLIF(EXCLUDED.sedol,           ''), figi_lookup.sedol),
        comp_figi       = EXCLUDED.comp_figi,
        shareclass_figi = EXCLUDED.shareclass_figi,
        sectype         = COALESCE(NULLIF(EXCLUDED.sectype,         ''), figi_lookup.sectype),
        sectype2        = COALESCE(NULLIF(EXCLUDED.sectype2,        ''), figi_lookup.sectype2),
        mkt_sector      = COALESCE(NULLIF(EXCLUDED.mkt_sector,      ''), figi_lookup.mkt_sector),
        update_at       = NOW()
"""


def fetch(
    id_pairs: list[tuple[str, str]],
    id_type: str,
    session: requests.Session,
    log: logging.Logger,
    job_extras: list[dict] | None = None,
) -> pd.DataFrame:
    """
    Call OpenFIGI /v3/mapping for a list of identifiers.

    id_pairs: list of (api_value, isin_tag) where:
      - api_value is the identifier sent to OpenFIGI (any id_type OpenFIGI accepts)
      - isin_tag is stored as 'isin' in result rows for home-exchange ranking

    job_extras: optional list of dicts (one per id_pair) merged into each OpenFIGI job,
      e.g. [{"exchCode": "US"}, {}, {"exchCode": "LN"}] to filter by exchange per row.

    Retries on HTTP 429. Returns a long DataFrame with FIGI_COLUMNS.
    """
    headers = {
        "Content-Type": "application/json",
        "X-OPENFIGI-APIKEY": OPENFIGI_API_KEY,
    }
    rows: list[dict] = []
    unmatched: list[str] = []
    batches = [id_pairs[i:i + OPENFIGI_BATCH] for i in range(0, len(id_pairs), OPENFIGI_BATCH)]
    log.info(f"  Calling OpenFIGI in {len(batches)} batch(es) of up to {OPENFIGI_BATCH} …")

    for batch_num, batch in enumerate(batches, 1):
        global_offset = (batch_num - 1) * OPENFIGI_BATCH
        jobs = []
        for local_i, (api_val, _) in enumerate(batch):
            job = {"idType": id_type, "idValue": api_val}
            if job_extras:
                extra = job_extras[global_offset + local_i]
                if extra:
                    job.update(extra)
            jobs.append(job)

        while True:
            resp = session.post(OPENFIGI_URL, json=jobs, headers=headers, timeout=30, verify=False)
            if resp.status_code == 429:
                wait = float(resp.headers.get("ratelimit-reset", 10)) + 0.5
                log.warning(f"  Rate-limited — waiting {wait:.1f}s …")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            break

        for (api_val, isin_tag), result in zip(batch, resp.json()):
            if "data" in result:
                for rec in result["data"]:
                    rows.append({"isin": isin_tag, **rec})
            else:
                msg = result.get("warning") or result.get("error") or "no data"
                unmatched.append(f"{api_val}: {msg}")

        log.info(f"  Batch {batch_num}/{len(batches)} done — {len(rows)} rows so far")

    if unmatched:
        log.warning(f"  {len(unmatched)} identifier(s) without data:")
        for line in unmatched:
            log.warning(f"    {line}")

    df = pd.DataFrame(rows)
    for col in FIGI_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA
    return df[FIGI_COLUMNS] if not df.empty else pd.DataFrame(columns=FIGI_COLUMNS)


def pick_representative(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reduce a long OpenFIGI mapping DataFrame to one representative row per ISIN.

    Within each ISIN group, prefers:
      1. Composite-level records (figi == compositeFIGI)
      2. Records on the ISIN's home exchange (via HOME_EXCH; earlier entries rank higher)
      3. Lowest FIGI alphabetically (deterministic tiebreaker)
    """
    if df.empty:
        return df.copy()

    work = df.copy()

    # Grouping key with fallbacks for null shareClassFIGI
    work["dedupe_key"] = (
        work["shareClassFIGI"]
        .fillna(work["compositeFIGI"])
        .fillna(work["figi"])
    )

    # Preference 1: composite-level record (figi == compositeFIGI)
    work["_is_composite"] = (
        work["figi"].notna()
        & work["compositeFIGI"].notna()
        & (work["figi"] == work["compositeFIGI"])
    ).astype(int)

    # Preference 2: home-exchange rank from the ISIN prefix
    def home_rank(row) -> int:
        isin = row.get("isin")
        if not isinstance(isin, str) or len(isin) < 2:
            return 99
        prefs = HOME_EXCH.get(isin[:2].upper(), [])
        try:
            return prefs.index(row["exchCode"])
        except (ValueError, TypeError):
            return 99

    work["_home_rank"] = work.apply(home_rank, axis=1)

    # Sort so the best row per ISIN comes first, then take the first per ISIN
    work = work.sort_values(
        by=["isin", "_is_composite", "_home_rank", "figi"],
        ascending=[True, False, True, True],
        kind="mergesort",
    )
    rep = (
        work.groupby("isin", as_index=False, sort=True)
        .first()
        .drop(columns=["_is_composite", "_home_rank"])
    )

    cols = ["dedupe_key"] + [c for c in df.columns if c in rep.columns]
    return rep[cols].reset_index(drop=True)


# ── CUSIP → ISIN (Luhn mod-10 check digit) ───────────────────────────────────

def _cusip_to_isin(cusip: str, country_code: str = "US") -> str:
    base = country_code + cusip.upper()
    digits = ""
    for ch in base:
        if ch.isdigit():
            digits += ch
        elif ch.isalpha():
            digits += str(ord(ch) - ord("A") + 10)
    digits += "0"
    total = 0
    for i, d in enumerate(reversed(digits)):
        n = int(d)
        if i % 2 == 1:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    check = str((10 - (total % 10)) % 10)
    return base + check


# ── Test harness ──────────────────────────────────────────────────────────────

_XLSX = EXCEL_DIR / "figi_utils.xlsx"
_INPUT_SHEET  = "Securities"
_OUTPUT_SHEET = "Results"


def test(
    xlsx_path: Path = _XLSX,
    input_sheet: str = _INPUT_SHEET,
    output_sheet: str = _OUTPUT_SHEET,
) -> None:
    """
    Read ISINs/CUSIPs from `input_sheet`, fetch FIGIs via OpenFIGI, and write
    one representative row per identifier to `output_sheet` in the same workbook.

    Input sheet must have columns: security_name, isin, cusip.
    Rows with an ISIN are looked up by ID_ISIN; ISIN-less rows use ID_CUSIP
    (CUSIP is converted to a derived ISIN for home-exchange ranking).
    """
    log = logging.getLogger("figi_utils.test")
    if not log.handlers:
        h = logging.StreamHandler(sys.stdout)
        h.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", "%H:%M:%S"))
        log.addHandler(h)
        log.setLevel(logging.DEBUG)

    df_in = pd.read_excel(xlsx_path, sheet_name=input_sheet, dtype=str).fillna("")
    log.info(f"Read {len(df_in)} row(s) from '{input_sheet}' in {xlsx_path.name}")

    isin_pairs:  list[tuple[str, str]] = []
    cusip_pairs: list[tuple[str, str]] = []

    for _, row in df_in.iterrows():
        isin  = row.get("isin",  "").strip()
        cusip = row.get("cusip", "").strip()
        if isin:
            isin_pairs.append((isin, isin))
        elif cusip and cusip.isalnum():
            cusip_pairs.append((cusip, _cusip_to_isin(cusip)))

    if not isin_pairs and not cusip_pairs:
        log.warning("No valid ISINs or CUSIPs found — nothing to fetch.")
        return

    frames: list[pd.DataFrame] = []
    with requests.Session() as session:
        if isin_pairs:
            log.info(f"Fetching {len(isin_pairs)} ISIN(s) …")
            frames.append(fetch(isin_pairs, "ID_ISIN", session, log))
        if cusip_pairs:
            log.info(f"Fetching {len(cusip_pairs)} CUSIP(s) …")
            frames.append(fetch(cusip_pairs, "ID_CUSIP", session, log))

    long_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=FIGI_COLUMNS)

    if long_df.empty:
        log.warning("No data returned from OpenFIGI.")
        return

    log.info(f"{len(long_df)} venue-level row(s) returned")
    rep_df = pick_representative(long_df)
    log.info(f"{len(rep_df)} representative row(s) after dedup")

    wb = openpyxl.load_workbook(xlsx_path)
    if output_sheet in wb.sheetnames:
        del wb[output_sheet]
    ws = wb.create_sheet(output_sheet)

    headers = list(rep_df.columns)
    ws.append(headers)
    for _, row in rep_df.iterrows():
        ws.append([None if pd.isna(v) else v for v in (row[h] for h in headers)])

    wb.save(xlsx_path)
    log.info(f"Results written to sheet '{output_sheet}' in {xlsx_path.name}")


if __name__ == "__main__":
    test()
