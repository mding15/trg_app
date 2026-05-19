"""
process_file_upload.py — Process positions from an xlsx file upload and insert into proc_positions.

Steps:
    1. Read the 'positions' tab from the xlsx file into a list of dicts.
    2. Batch-validate account_ids against the account table.
       Rows whose account_id is not found are skipped and logged.
    3. Batch-load security_xref cache for all ISINs and CUSIPs in rows
       where security_id is not already supplied by the file.
    4. Load cash_security_map (TRG CASH entries from security_info) for MF fallback.
    5. Per row:
       a. Skip if account_id is unknown.
       b. If security_id is populated in the file, use it directly (no lookup).
       c. Otherwise resolve via security_xref (ISIN first, then CUSIP).
          If still not found and asset_class='MF', try cash_security_map by currency.
          If still not found but at least one identifier exists, create a new security
          in security_info using asset_class and asset_type from the file.
          Rows with no identifier and no MF fallback are skipped and logged.
    6. Renumber position_id 1..n (only counting accepted rows).
    7. Archive to proc_positions_hist and replace same-date rows, scoped to
       feed_source='file_upload'. Uses the max as_of_date in the accepted rows as feed_date.
    8. Insert processed rows into proc_positions.

Usage:
    python process_file_upload.py
    python process_file_upload.py --file path/to/upload.xlsx
    python process_file_upload.py --account-id 1013
    python process_file_upload.py --file path/to/upload.xlsx --account-id 1013
"""
from __future__ import annotations

import logging
import os
import sys
from datetime import datetime

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from database2 import pg_connection

FEED_SOURCE = 'file_upload'
DEFAULT_FILE = os.path.join(os.path.dirname(__file__), 'test_data', 'file_upload.xlsx')
POSITIONS_SHEET = 'positions'


# ── logging setup ──────────────────────────────────────────────────────────────

def _setup_logger(feed_date, account_id=None) -> logging.Logger:
    log_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'log')
    os.makedirs(log_dir, exist_ok=True)
    account_suffix = f'_account{account_id}' if account_id is not None else ''
    log_file = os.path.join(
        log_dir,
        f'process_file_upload_{feed_date}{account_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log',
    )
    logger = logging.getLogger(f'process_file_upload_{feed_date}{account_suffix}')
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter('%(asctime)s  %(levelname)-8s  %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return logger


# ── file reading ───────────────────────────────────────────────────────────────

def _read_positions(filepath: str) -> list[dict]:
    """Read the 'positions' sheet from the xlsx file. Returns a list of row dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if POSITIONS_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{POSITIONS_SHEET}' not found in {filepath}. Available: {wb.sheetnames}")
    ws = wb[POSITIONS_SHEET]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)}
        # Skip completely blank rows
        if all(v is None for v in row.values()):
            continue
        rows.append(row)
    return rows


# ── batch cache loaders ────────────────────────────────────────────────────────

def _load_valid_account_ids(cur, raw_rows: list[dict], filter_account_id=None) -> set[int]:
    """
    Validate account_ids from the file against the account table.
    Returns the set of valid account_ids.
    If filter_account_id is given, only that one account_id is loaded.
    """
    if filter_account_id is not None:
        cur.execute('SELECT account_id FROM account WHERE account_id = %s', (filter_account_id,))
        return {row[0] for row in cur.fetchall()}

    ids = list({int(r['account_id']) for r in raw_rows if r.get('account_id') is not None})
    if not ids:
        return set()
    cur.execute('SELECT account_id FROM account WHERE account_id = ANY(%s)', (ids,))
    return {row[0] for row in cur.fetchall()}


def _load_security_cache(cur, raw_rows: list[dict]) -> dict[tuple, str]:
    """
    Batch-load security_xref rows for all ISINs and CUSIPs in rows where
    security_id is not already supplied (i.e. security_id is None in the file).
    Returns {('ISIN', isin): security_id, ('CUSIP', cusip): security_id, ...}.
    """
    needs_lookup = [r for r in raw_rows if not r.get('security_id')]
    isins  = list({r.get('isin')  or '' for r in needs_lookup if r.get('isin')})
    cusips = list({r.get('cusip') or '' for r in needs_lookup if r.get('cusip')})

    cache: dict[tuple, str] = {}

    if isins:
        cur.execute(
            'SELECT "REF_ID", "SecurityID" FROM security_xref WHERE "REF_TYPE" = %s AND "REF_ID" = ANY(%s)',
            ('ISIN', isins),
        )
        for ref_id, sec_id in cur.fetchall():
            cache[('ISIN', ref_id)] = sec_id

    if cusips:
        cur.execute(
            'SELECT "REF_ID", "SecurityID" FROM security_xref WHERE "REF_TYPE" = %s AND "REF_ID" = ANY(%s)',
            ('CUSIP', cusips),
        )
        for ref_id, sec_id in cur.fetchall():
            cache.setdefault(('CUSIP', ref_id), sec_id)

    return cache


def _load_cash_security_map(cur) -> dict[str, str]:
    """
    Load SecurityID for cash securities keyed by Currency.
    Returns {currency: SecurityID} for all cash securities with DataSource='TRG CASH'.
    Used to resolve MF (asset_class='MF') positions by currency when no ISIN/CUSIP is present.
    """
    cur.execute(
        """
        SELECT "SecurityID", "Currency"
        FROM security_info
        WHERE "AssetClass" = 'Cash' AND "DataSource" = 'TRG CASH'
        """
    )
    return {row[1]: row[0] for row in cur.fetchall()}


# ── security creation ──────────────────────────────────────────────────────────

def _create_security(
    cur,
    security_name: str,
    currency: str,
    asset_class: str | None,
    asset_type: str | None,
    logger: logging.Logger,
) -> str:
    """Insert a new row into security_info and return the generated SecurityID."""
    cur.execute(
        """
        INSERT INTO security_info
            ("SecurityName", "Currency", "AssetClass", "AssetType", "DataSource")
        VALUES (%s, %s, %s, %s, 'FILE_UPLOAD')
        RETURNING id
        """,
        (security_name, currency, asset_class, asset_type),
    )
    new_id = cur.fetchone()[0]
    security_id = f'T1{str(new_id).zfill(7)}'
    cur.execute(
        'UPDATE security_info SET "SecurityID" = %s WHERE id = %s',
        (security_id, new_id),
    )
    return security_id


def _add_xref_if_missing(cur, security_id: str, ref_type: str, ref_id: str) -> None:
    """Insert a security_xref row if ref_id is non-empty and not already present."""
    if not ref_id or not ref_id.strip():
        return
    cur.execute(
        'SELECT 1 FROM security_xref WHERE "REF_TYPE" = %s AND "REF_ID" = %s',
        (ref_type, ref_id),
    )
    if cur.fetchone():
        return
    cur.execute(
        """
        INSERT INTO security_xref ("REF_ID", "REF_TYPE", "SecurityID", "DataSource")
        VALUES (%s, %s, %s, 'FILE_UPLOAD')
        """,
        (ref_id, ref_type, security_id),
    )


def _get_or_create_security_id(
    cur,
    security_cache: dict[tuple, str],
    security_id_from_file: str | None,
    isin: str | None,
    cusip: str | None,
    ticker: str | None,
    security_name: str,
    currency: str,
    asset_class: str | None,
    asset_type: str | None,
    cash_security_map: dict[str, str],
    logger: logging.Logger,
) -> str | None:
    """
    Resolve SecurityID for a single row.

    1. If security_id is already in the file, return it directly (trusted).
    2. Try security_xref: ISIN first, then CUSIP.
    3. If both ISIN and CUSIP are missing and asset_class='MF', try cash_security_map by currency.
    4. If at least one identifier is present but not in the cache, create a new security
       using asset_class and asset_type from the file.
    5. Return None (row will be skipped) if no resolution is possible.
    """
    # ── Trust the file if security_id is already provided ─────────────────────
    if security_id_from_file:
        return security_id_from_file

    # ── Standard lookup: ISIN first, then CUSIP ───────────────────────────────
    if isin and ('ISIN', isin) in security_cache:
        return security_cache[('ISIN', isin)]

    if cusip and ('CUSIP', cusip) in security_cache:
        return security_cache[('CUSIP', cusip)]

    # ── Both ISIN and CUSIP missing ───────────────────────────────────────────
    if not isin and not cusip:
        if asset_class == 'MF':
            security_id = cash_security_map.get(currency)
            if security_id:
                return security_id

        logger.warning(
            f"Cannot resolve security_id: "
            f"security_name='{security_name}' currency='{currency}' asset_class='{asset_class}'"
        )
        return None

    # ── At least one identifier present but not in cache: create new security ─
    security_id = _create_security(cur, security_name or '', currency or '', asset_class, asset_type, logger)
    _add_xref_if_missing(cur, security_id, 'ISIN',   isin   or '')
    _add_xref_if_missing(cur, security_id, 'CUSIP',  cusip  or '')
    _add_xref_if_missing(cur, security_id, 'Ticker', ticker or '')

    if isin:
        security_cache[('ISIN', isin)]   = security_id
    if cusip:
        security_cache[('CUSIP', cusip)] = security_id

    logger.info(
        f"Created new security: name='{security_name}' isin='{isin}' cusip='{cusip}' → {security_id}"
    )
    return security_id


# ── archive helper ─────────────────────────────────────────────────────────────

def _archive_and_replace(cur, account_ids: list[int], feed_date, logger: logging.Logger) -> None:
    """
    For each account_id in account_ids, manage proc_positions rows before the new insert.
    Scoped to feed_source='file_upload'.

    - Rows with as_of_date < feed_date  → move to proc_positions_hist.
    - Rows with as_of_date = feed_date  → delete (will be replaced by the new insert).
    """
    ids = list(set(account_ids))

    cur.execute(
        """
        DELETE FROM proc_positions_hist h
        WHERE h.account_id = ANY(%s)
          AND h.feed_source = %s
          AND EXISTS (
              SELECT 1 FROM proc_positions p
              WHERE p.account_id = h.account_id
                AND p.as_of_date  = h.as_of_date
                AND p.feed_source = h.feed_source
                AND p.as_of_date  < %s
          )
        """,
        (ids, FEED_SOURCE, feed_date),
    )

    cur.execute(
        """
        INSERT INTO proc_positions_hist
            (as_of_date, account_id, position_id, security_id, security_name,
             isin, cusip, ticker, quantity, market_value, asset_class, currency,
             broker_account, broker, last_price, last_price_date, feed_source, insert_time, archived_at,
             total_cost)
        SELECT
            as_of_date, account_id, position_id, security_id, security_name,
            isin, cusip, ticker, quantity, market_value, asset_class, currency,
            broker_account, broker, last_price, last_price_date, feed_source, insert_time, NOW(),
            total_cost
        FROM proc_positions
        WHERE account_id = ANY(%s) AND as_of_date < %s AND feed_source = %s
        """,
        (ids, feed_date, FEED_SOURCE),
    )
    archived = cur.rowcount

    cur.execute(
        "DELETE FROM proc_positions WHERE account_id = ANY(%s) AND as_of_date < %s AND feed_source = %s",
        (ids, feed_date, FEED_SOURCE),
    )

    cur.execute(
        "DELETE FROM proc_positions WHERE account_id = ANY(%s) AND as_of_date = %s AND feed_source = %s",
        (ids, feed_date, FEED_SOURCE),
    )
    replaced = cur.rowcount

    if archived:
        logger.info(f"Archived {archived} rows to proc_positions_hist (as_of_date < {feed_date})")
    if replaced:
        logger.info(f"Deleted {replaced} existing proc_positions rows for as_of_date={feed_date} (to be replaced)")


# ── main ───────────────────────────────────────────────────────────────────────

def process_file_upload(filepath: str, account_id: int | None = None) -> int:
    """
    Read positions from an xlsx file and insert into proc_positions.
    If account_id is given, only rows for that account_id are processed.
    Returns the number of rows inserted.
    """
    # ── Step 1: read xlsx ───────────────────────────────────────────────────────
    raw_rows = _read_positions(filepath)
    if not raw_rows:
        print(f"No rows found in '{filepath}' sheet='{POSITIONS_SHEET}'")
        return 0

    # Derive feed_date from max as_of_date in file (used for archive step and logging)
    dates = [r['as_of_date'] for r in raw_rows if r.get('as_of_date')]
    feed_date = max(dates).date() if dates and hasattr(dates[0], 'date') else max(dates)

    logger = _setup_logger(feed_date, account_id)
    logger.info(f"=== Start processing file upload: file='{filepath}' feed_date={feed_date}"
                + (f" account_id={account_id}" if account_id is not None else "") + " ===")
    logger.info(f"Read {len(raw_rows)} rows from '{POSITIONS_SHEET}' tab")

    with pg_connection() as conn:
        with conn.cursor() as cur:
            # ── Step 2: validate account_ids ────────────────────────────────────
            valid_account_ids = _load_valid_account_ids(cur, raw_rows, account_id)
            logger.info(f"Validated {len(valid_account_ids)} account_id(s) against account table")

            # ── Step 3: batch-load security cache for rows missing security_id ──
            security_cache = _load_security_cache(cur, raw_rows)
            logger.info(f"Loaded {len(security_cache)} security_xref entries into cache")

            # ── Step 4: load cash security map for MF fallback ──────────────────
            cash_security_map = _load_cash_security_map(cur)
            logger.info(f"Loaded {len(cash_security_map)} TRG CASH entries into cash_security_map")

            missing_accounts: set = set()
            processed: list[dict] = []
            skipped = 0
            pos_idx = 1

            for r in raw_rows:
                row_account_id = r.get('account_id')
                if row_account_id is None:
                    skipped += 1
                    continue
                row_account_id = int(row_account_id)

                # ── Step 2 (per row): skip if account_id not valid ───────────────
                if row_account_id not in valid_account_ids:
                    if row_account_id not in missing_accounts:
                        missing_accounts.add(row_account_id)
                        logger.warning(
                            f"account_id={row_account_id} not found in account table — "
                            f"all rows for this account will be skipped"
                        )
                    skipped += 1
                    continue

                isin          = r.get('isin')          or None
                cusip         = r.get('cusip')         or None
                ticker        = r.get('ticker')        or None
                security_name = r.get('security_name') or ''
                currency      = r.get('currency')      or ''
                asset_class   = r.get('asset_class')   or None
                asset_type    = r.get('asset_type')    or None

                # ── Step 5: resolve security_id ──────────────────────────────────
                security_id = _get_or_create_security_id(
                    cur,
                    security_cache,
                    r.get('security_id') or None,
                    isin,
                    cusip,
                    ticker,
                    security_name,
                    currency,
                    asset_class,
                    asset_type,
                    cash_security_map,
                    logger,
                )
                if security_id is None:
                    skipped += 1
                    continue

                # Normalise as_of_date to a plain date
                as_of_date = r.get('as_of_date')
                if as_of_date is not None and hasattr(as_of_date, 'date'):
                    as_of_date = as_of_date.date()

                # ── Step 6: build processed row (position_id = 1..n) ─────────────
                processed.append({
                    'as_of_date':      as_of_date,
                    'account_id':      row_account_id,
                    'position_id':     str(pos_idx),
                    'security_id':     security_id,
                    'security_name':   security_name,
                    'isin':            isin,
                    'cusip':           cusip,
                    'ticker':          ticker,
                    'quantity':        r.get('quantity'),
                    'market_value':    r.get('market_value'),
                    'asset_class':     asset_class,
                    'currency':        currency,
                    'broker_account':  r.get('broker_account'),
                    'broker':          r.get('broker'),
                    'last_price':      r.get('last_price'),
                    'last_price_date': r.get('last_price_date'),
                    'feed_source':     FEED_SOURCE,
                    'total_cost':      r.get('total_cost'),
                })
                pos_idx += 1

            if skipped:
                logger.warning(f"Skipped {skipped} rows (unresolved account_id or security_id)")

            # ── Steps 7 & 8: archive, replace, insert ───────────────────────────
            if processed:
                account_ids = list({r['account_id'] for r in processed})
                _archive_and_replace(cur, account_ids, feed_date, logger)

                insert_sql = """
                    INSERT INTO proc_positions
                        (as_of_date, account_id, position_id, security_id, security_name,
                         isin, cusip, ticker, quantity, market_value, asset_class, currency,
                         broker_account, broker, last_price, last_price_date, feed_source,
                         total_cost)
                    VALUES
                        (%(as_of_date)s, %(account_id)s, %(position_id)s, %(security_id)s,
                         %(security_name)s, %(isin)s, %(cusip)s, %(ticker)s, %(quantity)s,
                         %(market_value)s, %(asset_class)s, %(currency)s, %(broker_account)s,
                         %(broker)s, %(last_price)s, %(last_price_date)s, %(feed_source)s,
                         %(total_cost)s)
                """
                cur.executemany(insert_sql, processed)

        conn.commit()

    n = len(processed)
    logger.info(f"Inserted {n} rows into proc_positions (feed_source='{FEED_SOURCE}', feed_date={feed_date})")
    logger.info("=== Done ===")
    return n


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Process file upload positions into proc_positions')
    parser.add_argument('--file', default=DEFAULT_FILE, metavar='PATH',
                        help=f'Path to the xlsx file (default: {DEFAULT_FILE})')
    parser.add_argument('--account-id', metavar='ACCOUNT_ID', type=int,
                        help='Limit processing to this account_id only')
    args = parser.parse_args()

    process_file_upload(args.file, args.account_id)
