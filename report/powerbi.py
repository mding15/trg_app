# -*- coding: utf-8 -*-
"""
Created on Sat Oct 21 11:54:17 2023

@author: mgdin

public:
    delete_report_from_db(port_id_list)
    generate_report(DATA)
    insert_results_to_db(results, username, report_description)
    read_results(rpt_folder)

"""

import pandas as pd
import numpy as np
import xlwings as xw
from pathlib import Path
from openpyxl import load_workbook

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

from trg_config import config
from engine import VaR_engine as eng
from security import security_info as si
from report.preprocess import preprocess
from mkt_data import mkt_data_info
from utils import data_utils, tools, var_utils, xl_utils, mkt_data
from database import db_utils, ms_sql_server
from report import portfolio as port

from api.logging_config import get_logger
logger = get_logger(__name__)

# global variables
def load_pe_stat():
    pe_stat = data_utils.load_stat('private_equity')
    if len(pe_stat.columns) == 0: # default value
        pe_stat = pd.DataFrame(columns=['Proxy', 'ProxyBeta','Liquidity Adjusted', 'Tail_Shock'])
    return pe_stat


###############################################################################
# main function
def generate_report(DATA):

    try:
        # check input data
        check_input(DATA)
        
        # pre-process, calc peforance, sharp ratio, etc..
        preprocess(DATA)
    
        # generate PowerBI data
        results = gen_pbi_data_master(DATA)    
        
    except Exception as e:
        # print(e)
        raise e

    return results

def generate_report2(DATA):
    results = generate_report(DATA)
    
    # consolidate results to one book
    book = consolidate_books(results)
    
    # rename Positions to DimPositions
    book = rename_tab(book)

    return book
    
    
###############################################################################
def gen_pbi_data_master(DATA):

    results = {}
    
    results['dm_fact_d_Positions']     = Fact_Positions_Master(DATA)
    results['dm_fact_d_MgPositions']   = Fact_MgPositions_Master(DATA)
    results['dm_fact_d_Parameters']    = Fact_Parameters(DATA)
    results['dm_d_concentration_2']        = Fact_Limits(DATA)
    results['dm_fact_d_AggTable']      = Fact_AggTable_Master(DATA)
    results['dm_d_Positions']          = Positions_Master(DATA)
    results['dm_d_Unknown_Position']   = Positions_Exception(DATA)
    results['dm_d_Top_Risks_and_Hedges']  = Top_Risks_and_Hedges_Master(DATA)
    results['dm_d_Lims_Concentration']  = Lims_Concentration_Master(DATA)
    results['dm_d_concentration']       = d_concentration_Master(DATA)
    results['dm_d_Lims_VaR']            = Lims_VaR_Master(DATA)
    results['dm_d_Lims_Vol']            = Lims_Vol_Master(DATA)
    results['dm_d_TS']                  = TS_Master(DATA)
    results['dm_d_TS_Vol_VaR']        = DATA['BackTest'] # use table_name as key
    results['dm_port_consolidated']   = port.get_port_consolidated(DATA)
    results['dm_port_hist_value']     = port.get_port_hist_value(DATA)
    results['dm_port_return']         = port.get_port_returns(DATA)
    
    results.update(DimClasses_Master(DATA))
    
    return results    

def Book(df):
    return {'Tab': df}


def DimClasses_Master(DATA):
    results = {}
    
    file_path = config['REFD_DIR'] / 'DimClasses.xlsx'
    workbook = load_workbook(file_path)
    for tab in workbook.sheetnames:
        # print(f'dm_dim_d_{tab}')
        df = pd.read_excel(file_path, sheet_name=tab)
        results[f'dm_dim_d_{tab}'] = df
    
    # for f in list(ref_dir.glob("Dim*.xlsx")):
    #     print(f.stem)
    #     dfs = pd.read_excel(f, sheet_name=None)
    #     tab = list(dfs.keys())[0]
    #     book = {tab: dfs[tab]}
    #     results[f.stem] = book
        
    return results
    # names = ['DimClasses', 'DimCountry', 'DimCurrency', 'DimGICS', 'DimIndustry','DimRegion','DimSC1','DimSC2','DimSC3','DimSectors']
    # tabs  = ['DimClasses', 'DimCountry', 'DimCurrency', 'DimGICS', 'DimIndustry','DimRegion','DimSC1','DimSC2','DimSC3','DimSector']

    # for i in range(len(names)):
    #     name = names[i]    
    #     tab  = tabs[i]
    #     filename = config['REFD_DIR'] / f'{name}_Master.xlsx'    
    #     #print(name, tab, filename)
        
    #     df = pd.read_excel(filename, sheet_name=tab)
    #     book = {tab: df}
    #     results[f'{name}_Master'] = book
    
    
def Lims_Concentration_Master(DATA):
    tab = 'Lims Concentration'
    
    filename = config['REFD_DIR'] / 'Lims_Concentration_Master.xlsx'
    df = pd.read_excel(filename, sheet_name=tab)
    df.columns = ['Concentration'] + list(df.columns[1:])
    df.rename(columns={'Low ':'Low'}, inplace=True)
    
    return df

def d_concentration_Master(DATA):
    port_id = DATA['port_id']

    # read limits from database    
    query = f"""
    select lc.* from portfolio_info pi, limit_concentration lc   
    where pi.port_group_id = lc.port_group_id 
    and pi.port_id = {port_id}
    """
    df = db_utils.get_sql_df(query)
    df.rename(columns={'category': 'Category', 'limit_value': 'Limit'}, inplace=True)
    df.drop(columns=['port_group_id'], inplace=True)
    
    
    return df

def test_d_concentration_Master(DATA):
    book = d_concentration_Master(DATA)

def Lims_VaR_Master(DATA):
    port_id = DATA['port_id']

    # read limits from database    
    query = f"""
    select lv.* from portfolio_info pi, limit_var lv  
    where pi.port_group_id = lv.port_group_id 
    and pi.port_id = {port_id} and lv.risk_type = 'VaR'

    """
    df = db_utils.get_sql_df(query)
    df.rename(columns={'low': 'Low', 'mid1': 'Mid1', 'mid2': 'Mid2', 'high': 'High'}, inplace=True)
    df.drop(columns=['port_group_id'], inplace=True)
    
    # read limit from file
    # filename = config['REFD_DIR'] / 'Lims_VaR_Master.xlsx'
    # df = pd.read_excel(filename, sheet_name=tab)
    # df.columns = ['risk_type'] + list(df.columns[1:])
    # df.rename(columns={'Low ':'Low'}, inplace=True)
    
    return df

def Lims_Vol_Master(DATA):
    port_id = DATA['port_id']

    # read limits from database    
    query = f"""
    select lv.* from portfolio_info pi, limit_var lv  
    where pi.port_group_id = lv.port_group_id 
    and pi.port_id = {port_id} and lv.risk_type = 'Vol'

    """
    df = db_utils.get_sql_df(query)
    df.rename(columns={'low': 'Low', 'mid1': 'Mid1', 'mid2': 'Mid2', 'high': 'High'}, inplace=True)
    df.drop(columns=['port_group_id'], inplace=True)
    
    # read limit from file
    # filename = config['REFD_DIR'] / 'Lims_Vol_Master.xlsx'
    # df = pd.read_excel(filename, sheet_name=tab)
    # df.columns = ['risk_type'] + list(df.columns[1:])
    # df.rename(columns={'Low ':'Low'}, inplace=True)
    
    return df

def TS_Master(DATA):
    tab = 'TS'
    filename = config['REFD_DIR'] / 'TS.xlsx'
    
    df = pd.read_excel(filename, sheet_name=tab)
    df = df.rename(columns={'Cum Return': 'Cum_Return', 'Return': 'Return_Value'})
    
    df['Return_Value'] = pd.to_numeric(df['Return_Value'], errors='coerce')
    df['Cum_Return'] = pd.to_numeric(df['Cum_Return'], errors='coerce')
    df['VaR'] = pd.to_numeric(df['VaR'], errors='coerce')
    
    df['Return_Value'] = df['Return_Value'].round(4)
    df['Cum_Return'] = df['Cum_Return'].round(4)
    df['VaR'] = df['VaR'].round(4)
    df['Return_Value'].fillna(0,inplace=True)
    df['Cum_Return'].fillna(0,inplace=True)  
    
    book = {tab: df}
    
    return df

def TS_Vol_VaR(DATA):
    df = DATA['BackTest']
    book = {'TS_Vol_VaR': df}
    
    return df

###############################################################################
# position exception
def Positions_Exception(DATA):
    df = DATA['position_exception']
    df = df[['ID','SecurityID','SecurityName','ISIN','CUSIP','Ticker','Quantity','MarketValue','userAssetClass','userCurrency']].copy()
    # df = df.rename(columns={'userAssetClass': 'AssetClass', 'userCurrency': 'Currency'})
    book = {'Positions_Exception': df}

    return df

###############################################################################
# Master version
    
def Fact_Positions_Master(DATA):
    positions = DATA['Positions']

    total_mv = positions['MarketValue'].sum()
    
    pos = pd.DataFrame(index=positions.index)
    pos['Security_ID']     = positions['SecurityID']
    pos['Security_Name']   = positions['SecurityName']
    pos['Market_Value']    = positions['MarketValue']
    pos['MV_Weight']       = positions['MarketValue'] / total_mv
    pos['Vol']             = positions['Volatility'].fillna(0)
    pos['VaR']             = positions['tVaR'].fillna(0) 
    
    # expected return
    pos['Exp_Ret'] = positions['ExpectedReturn']
                                   
    # Sharpe Ratio to Volatility
    pos['SR_Vol'] = positions['SR Vol']
    
    # Sharpe Ratio to VaR
    pos['SR_VaR'] = positions['SR tVaR']
    
    # asset classes
    pos['Class'] = positions['Class']
    pos['SC1'] = positions['SC1']
    pos['SC2'] = positions['SC2']
    
    # currency/country/region/industry
    pos['Currency']  = positions['Currency']
    pos['Country']   = positions['Country']
    pos['Region']    = positions['Region']
    pos['Industry']  = positions['Industry']
    
    # Sensitivity
    pos['IR_Duration']      = positions['Duration']
    pos['IR_Convexity']     = positions['Convexity']
    pos['Spread_Duration']  = positions['SpreadDuration']
    pos['Spread_Convexity'] = positions['SpreadConvexity']

    pos['Delta']         = positions['DELTA']
    pos['Gamma']         = positions['GAMMA']
    pos['Vega']          = positions['VEGA']
    pos['Implied_Vol']   = positions['IV']
    pos['PD'] = None
    
    labels = {
        'Delta'	: 'DELTA',
        'Gamma' : 'GAMMA',
        'Vega' : 'VOL',
        'IR_Duration': 'IR',
        'IR_Convexity' : 'IR_CV',
        'Spread_Duration': 'SPREAD',
        'Spread_Convexity' : 'SPREAD_CV',
        'Default': 'DEFAULT'}
    
    for label, cat in labels.items():
        #print(label, cat)
        col = cat + " VaR"
        if col in positions:
            pos[label + '_VaR'] = positions[col]
        else:
            pos[label + '_VaR'] = 0
        
    
    pos['Rating'] = None
    pos['Skewness'] = None
    pos['Tail_Fatness'] = None
    
    # proxy
    pe_stat = load_pe_stat()
    pos['Proxy'] = tools.df_series_merge(pos, pe_stat['Proxy'], key='Security_ID')
    pos['Proxy_Corr'] = tools.df_series_merge(pos, pe_stat['ProxyCorr'], key='Security_ID')
    # pos['Proxy Beta'] = tools.df_series_merge(pos, pe_stat['ProxyBeta'], key='Security_ID')
    pos['Liquidity_Adjusted'] = tools.df_series_merge(pos, pe_stat['Liquidity Adjusted'], key='Security_ID')
    # pos['Tail Shock'] = tools.df_series_merge(pos, pe_stat['Tail_Shock'], key='Security_ID')
    
    # add ticker
    pos['Ticker'] = positions['Ticker']
    # pos.loc[pos['Ticker'].isna()]['Ticker'] = positions.loc[pos['Ticker'].isna()]['ISIN']
    # pos.loc[pos['Ticker'].isna()]['Ticker'] = positions.loc[pos['Ticker'].isna()]['SecurityID']

    # max length
    pos['Security_Name'] = pos['Security_Name'].astype(str).str.slice(0, 100)
    pos['Class'] = pos['Class'].astype(str).str.slice(0, 50)
    pos['SC1'] = pos['SC1'].astype(str).str.slice(0, 50)
    pos['SC2'] = pos['SC2'].astype(str).str.slice(0, 50)
    pos['Currency'] = pos['Currency'].astype(str).str.slice(0, 50)
    pos['Country'] = pos['Country'].astype(str).str.slice(0, 50)
    pos['Region'] = pos['Region'].astype(str).str.slice(0, 50)
    pos['Industry'] = pos['Industry'].astype(str).str.slice(0, 50)
    pos['Rating'] = pos['Rating'].astype(str).str.slice(0, 50)
    pos['Proxy'] = pos['Proxy'].astype(str).str.slice(0, 50)
    pos['Ticker'] = pos['Ticker'].astype(str).str.slice(0, 50)

    book = {'Fact_Positions': pos}
    DATA['Fact_Positions'] = pos # save for later use

    return pos


# Marginal VaR
def Fact_MgPositions_Master(DATA):

    book = {}
    # Marginal Position
    positions = DATA['Positions'].set_index('pos_id')

    mg_pos = positions[['SecurityID','MarketValue', 'Weight', 'Class', 'SC1', 'SC2', 'Currency', 'Country', 'Region', 'Industry', 'Ticker']].copy()
    mg_pos = mg_pos.rename(columns={'SecurityID':'Security_ID', 'MarketValue':'Market_Value', 'Weight':'MV_Weight'})
    
    mg_pos['VaR'] = positions['tVaR']
    mg_pos['Mg_VaR_to_Top_D']   = positions['Marginal_tVaR']
    mg_pos['Mg_VaR_to_Top_P']   = mg_pos['Mg_VaR_to_Top_D'] / positions['Marginal_tVaR'].sum()
    mg_pos['Annual_Volatility'] = positions['Volatility']
    mg_pos['Mg_Vol_to_Top_P']   = positions['Marginal_Vol'] / positions['Marginal_Vol'].sum()
    
    columns = ['Security_ID', 'VaR', 'Mg_VaR_to_Top_D', 'Mg_VaR_to_Top_P', 'Annual_Volatility', 'Mg_Vol_to_Top_P', 'MV_Weight', 'Ticker']
    book['Fact_MgPositions'] = mg_pos[columns]
    
    return mg_pos[columns]

def Fact_AggTable_Master(DATA):
    
    book = {}

    # total portfolio risk    
    tot = DATA['TotalVaR']
    T = DATA['Parameters']['HorizonDays']
    
    # benchmark risk    
    bm = DATA['BechmarkRisk']
    
    db_risk = pd.DataFrame(columns=['Type','Vol_P','VaR_P','Exp_Ret','SR_Vol','SR_VaR','VaR_D'])
    db_risk.loc[0] = ['Portfolio', tot['Volatility'], tot['tVaR%'], tot['ExpectedReturn'], tot['SR Vol'], tot['SR tVaR'], tot['Marginal_tVaR']]
    db_risk.loc[1] = ['Benchmark', bm['Vol'], bm['VaR%'], bm['ExpRet'], bm['SR Vol'], bm['SR VaR'], np.nan]

    db_risk['VaR_a'] = db_risk['VaR_P'] * np.sqrt(252/T)
    # risk summary
    book['Fact_Agg Table'] = db_risk

    return db_risk

def Positions_Master(DATA):
    book = {}
    
    pos = DATA['Fact_Positions']
    df = pos[['Security_ID',  'Class', 'SC1', 'SC2', 'Currency', 'Country', 'Region', 'Industry']].copy()

    book['Positions'] = df
    return df
    

def Top_Risks_and_Hedges_Master(DATA):
    book = {}
    positions = DATA['Positions']
    
    df = positions[['SecurityName', 'Ticker', 'Weight', 'Volatility', 'Marginal_tVaR', 'ExpectedReturn', 'Marginal_Vol', 'tVaR Contribution','SecurityID']].copy()
    df['Return_Contrib'] = df['ExpectedReturn'] * df['Weight']
    df = df.rename(columns={'SecurityName':'Security_Name', 'Weight': 'MV_Weight', 'Marginal_tVaR': 'VaR', 'ExpectedReturn': 'Exp_Ret', 'Marginal_Vol': 'Vol_Contrib', 'tVaR Contribution': 'VaR_Contrib', 'SecurityID':'Security_ID'})
    
    df  = df.fillna(0)
    top_hedge = df[df['VaR']<0].copy()
    top_risk  = df[df['VaR']>0].copy()
    
    top_risk = top_risk.sort_values(by=['VaR'], ascending=False)
    top_hedge = top_hedge.sort_values(by=['VaR'], ascending=True)


    top_risk_and_hedge = pd.concat([top_risk.iloc[:10], top_hedge.iloc[:10]], ignore_index=True)
    book['Top Risks and Hedges'] = top_risk_and_hedge
    return top_risk_and_hedge
        
    
###############################################################################
def DimClasses(DATA):
    
    positions = DATA['Positions']
    data = positions.rename(columns = {'SecurityID': 'Security_ID'})
    data = data[['Class', 'SC1', 'SC2', 'Currency', 'Country', 'Region', 'Sector', 'Industry']]

    book = {}

    # Class
    df = data[['Class']].drop_duplicates()
    book['Class'] = df

    # SC1
    df = data[['Class', 'SC1']].drop_duplicates()
    book['SC1'] = df

    # SC2
    df = data[['Class', 'SC1', 'SC2']].drop_duplicates()
    book['SC2'] = df

    # Currency
    df = data[['Currency']].drop_duplicates()
    book['Currency'] = df

    # Region
    df = data[['Region']].drop_duplicates()
    book['Region'] = df

    # Country
    df = data[['Country', 'Region']].drop_duplicates()
    book['Country'] = df

    # Sector
    df = data[['Sector']].drop_duplicates()
    book['Sector'] = df

    # GICS
    df = data[['Sector', 'Industry']].drop_duplicates()
    book['GICS'] = df
    
    return book
    
def DimPositions(DATA):
    positions = DATA['Positions']
    dimPositions = positions.rename(columns = {'SecurityID': 'Security_ID', 'UserGroup1': 'Class', 'UserGroup2': 'SC1', 'UserGroup3': 'SC2'})
    dimPositions = dimPositions[['Security_ID', 'Class', 'SC1', 'SC2', 'Currency', 'Country', 'Region', 'Industry']]
    
    book = {'DimPositions': dimPositions}
    return book
    

def Fact_Parameters(DATA):
    parameters = DATA['Parameters']
    positions  = DATA['Positions']
    
    df = pd.DataFrame(index=[0])
    df['Portfolio_Name'] = parameters['PortfolioName']
    df['Portfolio_Size_mm'] = positions['MarketValue'].sum() / 1e6
    df['Report_Date'] = parameters['ReportDate']
    df['As_of_Date'] = parameters['AsofDate']
    df['Var_Vol_Window'] = parameters['RiskHorizon']
    df['Return_Frequency'] = 'Daily'
    df['Tail_Measure'] = '95% TailVaR'
    df['Benchmarks'] = parameters['Benchmark']
    df['Expected_Returns'] = 'Upload'
    
    if parameters.get('TrackedOvertime') == 'Y':
        df['Account_Name'] = parameters['PortfolioName']
        df['account_id'] = parameters.get('account_id')
    else:
        df['Account_Name'] = None
        df['account_id'] = None
        
    book = {'Fact_Parameters': df}
    return df

def Fact_Limits(DATA):
    """Process limit data for SQL Server insertion"""
    limit = DATA.get('limit', {})
    
    if not limit or limit == {}:
        # Return empty DataFrame with correct structure if no limit data
        df = pd.DataFrame(columns=['port_id', 'limit_category', 'limit_value'])
        return df
    
    # Convert limit dict to DataFrame
    limit_data = []
    for limit_category, limit_value in limit.items():
        if limit_category != 'port_id':  # Skip port_id field
            limit_data.append({
                'port_id': limit.get('port_id'),
                'limit_category': limit_category,
                'limit_value': limit_value
            })
    
    if limit_data:
        df = pd.DataFrame(limit_data)
    else:
        df = pd.DataFrame(columns=['port_id', 'limit_category', 'limit_value'])
    
    return df

def Fact_Positions(DATA):
    
    positions = DATA['Positions']
    
    rpt_pos = positions[['SecurityID', 'SecurityName', 'MarketValue']].copy()
    rpt_pos = rpt_pos.rename(columns={'SecurityID': 'Security_ID', 'SecurityName': 'Security Name', 'MarketValue': 'Market Value'})

    total_mv = rpt_pos['Market Value'].sum()
    rpt_pos['MV Weight'] = rpt_pos['Market Value'] / total_mv
    
    rpt_pos['Vol'] = positions['Volatility']
    rpt_pos['VaR'] = positions['tVaR'] 
    
    rpt_pos['VaR'].fillna(0, inplace=True)
    rpt_pos['Vol'].fillna(0, inplace=True)
    
    rpt_pos['Exp Ret'] = positions['ExpectedReturn']
    
    # Sharpe Ratio to Volatility
    rpt_pos['SR Vol'] = positions['SR Vol']
    
    # Sharpe Ratio to VaR
    rpt_pos['SR VaR'] = positions['SR tVaR']
    
    # asset classes
    rpt_pos['Class'] = positions['Class']
    rpt_pos['SC1'] = positions['SC1']
    rpt_pos['SC2'] = positions['SC2']
    
    # currency/country/region/industry
    rpt_pos['Currency'] = positions['Currency']
    rpt_pos['Country'] = positions['Country']
    rpt_pos['Region'] = positions['Region']
    rpt_pos['Industry'] = positions['Industry']
    
    # Sensitivity

    rpt_pos['IR Duration'] = positions['Duration']
    rpt_pos['IR Convexity'] = positions['Convexity']
    rpt_pos['Spread Duration'] = positions['SpreadDuration']
    rpt_pos['Spread Convexity'] = positions['SpreadConvexity']

    rpt_pos['Delta'] = positions['DELTA']
    rpt_pos['Gamma'] = positions['GAMMA']
    rpt_pos['Vega'] = positions['VEGA']
    rpt_pos['Implied Vol'] = positions['IV']
    rpt_pos['PD'] = None
    
    labels = {
        'Delta'	: 'DELTA',
        'Gamma' : 'GAMMA',
        'Vega' : 'VOL',
        'IR Duration': 'IR',
        'IR Convexity' : 'IR_CV',
        'Spread Duration': 'SPREAD',
        'Spread Convexity' : 'SPREAD_CV',
        'Default': 'DEFAULT'}
    
    for label, cat in labels.items():
        #print(label, cat)
        col = cat + " VaR"
        if col in positions:
            rpt_pos[label + ' VaR'] = positions[col]
        else:
            rpt_pos[label + ' VaR'] = 0
        
    
    rpt_pos['Rating'] = None
    rpt_pos['Skewness'] = None
    rpt_pos['Tail Fatness'] = None
    
    # proxy
    pe_stat = load_pe_stat()
    rpt_pos['Proxy'] = tools.df_series_merge(rpt_pos, pe_stat['Proxy'], key='Security_ID')
    rpt_pos['Proxy Corr'] = tools.df_series_merge(rpt_pos, pe_stat['ProxyCorr'], key='Security_ID')
    rpt_pos['Liquidity Adjusted'] = tools.df_series_merge(rpt_pos, pe_stat['Liquidity Adjusted'], key='Security_ID')
    
    # add ticker
    rpt_pos['Ticker'] = positions['Ticker']
    rpt_pos.loc[rpt_pos['Ticker'].isna()]['Ticker'] = positions.loc[rpt_pos['Ticker'].isna()]['ISIN']
    rpt_pos.loc[rpt_pos['Ticker'].isna()]['Ticker'] = positions.loc[rpt_pos['Ticker'].isna()]['SecurityID']

    book = {'Fact_Positions': rpt_pos}
    DATA['Fact_Positions'] = rpt_pos # save for later use

    return book


# Marginal VaR
def Fact_MgPositions(DATA):

    book = {}
    # Marginal Position
    positions = DATA['Positions'].set_index('pos_id')
    T = DATA['Parameters']['HorizonDays']
    PnL = DATA['PnL']

    mg_pos = positions[['SecurityID','MarketValue', 'Weight', 'Class', 'SC1', 'SC2', 'Currency', 'Country', 'Region', 'Industry', 'Ticker']].copy()
    mg_pos = mg_pos.rename(columns={'SecurityID':'Security_ID', 'MarketValue':'Market Value', 'Weight':'MV Weight'})
    
    mg_pos['VaR'] = positions['tVaR']
    mg_pos['Mg VaR to Top-D']   = positions['Marginal_tVaR']
    mg_pos['Mg VaR to Top-P']   = mg_pos['Mg VaR to Top-D'] / positions['Marginal_tVaR'].sum()
    mg_pos['Annual Volatility'] = positions['Volatility']
    mg_pos['Mg Vol to Top-P']   = positions['Marginal_Vol'] / positions['Marginal_Vol'].sum()
    
    columns = ['Security_ID', 'VaR', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Annual Volatility', 'Mg Vol to Top-P', 'MV Weight', 'Ticker']
    book['Fact_MgPositions'] = mg_pos[columns]
    
    # Construct hierarchy P/L
    hierarchy = ['Class', 'SC1', 'SC2', 'SecurityID']
    for col in hierarchy:
        positions[col].fillna('', inplace=True)
    hier_keys = positions.loc[PnL.columns, hierarchy]
    paths = hier_keys.apply(lambda x: '|'.join(x), axis=1).to_list()
    pos_pl = pd.DataFrame(PnL.values, columns=paths)

    for col in ['Class', 'SC1', 'SC2']:
        mg_pos[col].fillna('', inplace=True)

    # Marginal SC2
    mg_var = mg_pos.groupby(by=['Class', 'SC1', 'SC2'])[['Market Value', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Mg Vol to Top-P']].sum().reset_index()
    mg_var['Path'] = mg_var[['Class', 'SC1', 'SC2']].apply(lambda x: '|'.join(x), axis=1)
    mg_var = mg_var.set_index('Path')

    agg_pl = calc_agg_pl(pos_pl)
    mg_var['VaR'] = var_utils.calc_tVaR(agg_pl) * np.sqrt(T)
    mg_var['Annual Volatility'] = agg_pl.std() /mg_var['Market Value'] * np.sqrt(252)
    mg_var = mg_var[['Class', 'SC1', 'SC2', 'VaR', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Annual Volatility', 'Mg Vol to Top-P']]
    
    book['SC2'] = mg_var
    
    # Marginal SC1
    mg_var = mg_pos.groupby(['Class', 'SC1'])[['Market Value', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Mg Vol to Top-P']].sum().reset_index()
    mg_var['Path'] = mg_var[['Class', 'SC1']].apply(lambda x: '|'.join(x), axis=1)
    mg_var = mg_var.set_index('Path')
    
    agg_pl = calc_agg_pl(agg_pl)
    mg_var['VaR'] = var_utils.calc_tVaR(agg_pl) * np.sqrt(T)
    mg_var['Annual Volatility'] = agg_pl.std() /mg_var['Market Value'] * np.sqrt(252)
    mg_var = mg_var[['Class', 'SC1', 'VaR', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Annual Volatility', 'Mg Vol to Top-P']]
    
    book['SC1'] = mg_var
    
    # Marginal Class
    mg_var = mg_pos.groupby(['Class'])[['Market Value', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Mg Vol to Top-P']].sum().reset_index()
    mg_var['Path'] = mg_var[['Class']].apply(lambda x: '|'.join(x), axis=1)
    mg_var = mg_var.set_index('Path')
    
    agg_pl = calc_agg_pl(agg_pl)
    mg_var['VaR'] = var_utils.calc_tVaR(agg_pl) * np.sqrt(T)
    mg_var['Annual Volatility'] = agg_pl.std() /mg_var['Market Value'] * np.sqrt(252)
    mg_var = mg_var[['Class', 'VaR', 'Mg VaR to Top-D', 'Mg VaR to Top-P', 'Annual Volatility', 'Mg Vol to Top-P']]
    book['Class'] = mg_var
    
    return book



def Fact_AggTables(DATA):
    
    book = {}

    # total portfolio risk    
    tot = DATA['TotalVaR']

    # benchmark risk    
    bm = DATA['BechmarkRisk']
    
    db_risk = pd.DataFrame(columns=['Type','Vol_P','VaR_P','Exp Ret','SR Vol','SR VaR','VaR_D'])
    db_risk.loc[0] = ['Portfolio', tot['Volatility'], tot['tVaR%'], tot['ExpectedReturn'], tot['SR Vol'], tot['SR tVaR'], tot['Marginal_tVaR']]
    db_risk.loc[1] = ['Benchmark', bm['Vol'], bm['VaR%'], bm['ExpRet'], bm['SR Vol'], bm['SR VaR'], np.nan]

    # risk summary
    book['DB Risk'] = db_risk

    # top risk    
    top_risk, top_hedge = calc_top_risk(DATA)
    book['Top Risks'] = top_risk
    book['Top Hedges'] = top_hedge

    return book

def Fact_Bmk(DATA):

    # returns_1Y = DATA['Return_1Y']
    # book = {'Fact_Bmk': returns_1Y}
    book = {}    
    return book

   
#########################################################################
#
# tools
#


# aggregate pnl
def calc_agg_pl(pos_pl):
    parents = []
    for x in pos_pl.columns:
        keys = x.split('|')
        parents.append(['|'.join(keys[:-1]), x])
    parents = pd.DataFrame(parents, columns=['parent', 'path'])
    
    agg_pl = pd.DataFrame()
    for p in parents['parent'].unique():
        paths = parents[parents['parent']==p]['path']
        pl = pos_pl[paths]
        pl_sum = pl.sum(axis=1).rename(p)
        agg_pl = pd.concat([agg_pl, pl_sum], axis=1)

    return agg_pl

# top risk

def calc_top_risk(DATA):

    positions = DATA['Positions']
    
    df = positions[['SecurityName', 'Ticker', 'Weight', 'Volatility', 'Marginal_tVaR', 'ExpectedReturn', 'Marginal_Vol', 'VaR Contribution']].copy()
    df['Return Contrib'] = df['ExpectedReturn'] * df['Weight']
    df = df.rename(columns={'SecurityName':'Security_ID', 'Weight': 'MV Weight', 'Marginal_tVaR': 'VaR', 'ExpectedReturn': 'Exp Ret', 'Marginal_Vol': 'Vol Contrib', 'VaR Contribution': 'VaR Contrib'})
    
    
    df  = df.fillna(0)
    top_hedge = df[df['VaR']<0].copy()
    top_risk  = df[df['VaR']>0].copy()
    
    top_risk = top_risk.sort_values(by=['VaR'], ascending=False)
    top_hedge = top_hedge.sort_values(by=['VaR'], ascending=True)
    
    return top_risk, top_hedge

#########################################################################
#
# generate pbi data
#
def gen_pbi_data(DATA):

    results = {}
    
    results['DimClasses']       = DimClasses(DATA)
    results['DimPositions']     = DimPositions(DATA)
    results['Fact_Parameters']  = Fact_Parameters(DATA)
    results['Fact_Positions']   = Fact_Positions(DATA)
    results['Fact_MgPositions'] = Fact_MgPositions(DATA)
    results['Fact_AggTables']   = Fact_AggTables(DATA)
    results['Fact_Bmk']         = Fact_Bmk(DATA)

    return results    

    
# pbi output folder
def pbi_folder(params):
    
    #params = DATA['Parameters']
    client_id = params['ClientID']
    folder_name = params['PortfolioID']

    output_folder = config['DATA_DIR'] / 'powerbi' / client_id / folder_name
    output_folder.mkdir(parents=True, exist_ok=True)
    
    return output_folder
        
# write pbi files
def write_results(results, output_folder):
            
    for name in results:
        book = results[name]
        outfile = output_folder / f'{name}.xlsx'

        with pd.ExcelWriter(outfile) as writer:
            for tab in book:
                book[tab].to_excel(writer, sheet_name=tab, index=False)
    
        print('saved file:', outfile)

# read results from a given folder that contains all Excel files
def read_results(rpt_folder):
    # rpt_folder = Path(r'C:\Users\mgdin\Downloads\TailRiskGlobal\PowerBI\Model_1')
    
    results = {}
    
    for file_path in rpt_folder.glob('*.xlsx'):
        filename = file_path.stem
        book = {}
        results[filename] = book
        
        dfs = pd.read_excel(file_path, sheet_name=None)
        for tab, df in dfs.items():
            print(f'{filename}.{tab}')
            book[tab] = df
            
    return results

################################################## 
# write results to xl
# excel_file = config['SRC_DIR'] / 'test_data'  / 'Model-1.xlsx'

def write_results_xl(results, excel_file):
    
    # consolidate results to one book
    # book = consolidate_books(results)
    # book = consolidate_books_single_tab(results)
    
    # rename Positions to DimPositions
    book = rename_tab(results)
    
    # write book to excel_file    
    xl_utils.write_book_to_xl(book, excel_file)
    

# excel_file = file_path
# excel_file = r'C:/DATA/trgapp_data/clients/1013/13/5034.pbi.xlsx'
def read_results_xl(excel_file):
    
    return xl_utils.read_book_xl(excel_file)
    
# consolidate to one book
def consolidate_books(results):
    one_book = {}
    for name in results:
        book = results[name]
        for tab in book:
            if tab in one_book:
                raise Exception(f'found repeated tab name: {tab}')
            one_book[tab] = book[tab]
    return one_book
    
# consolidate to one book, each book has a single tab, so use book name
def consolidate_books_single_tab(results):
    one_book = {}
    for name in results:
        book = results[name]
        for tab in book:
            one_book[name] = book[tab]
    return one_book

# rename Positions to DimPositions            
def rename_tab(book):
    if 'Positions' in book:
        book['DimPositions'] = book.pop('Positions')

    return book

##################################################        
# on windows, use xlwings
def write_results_xl2(results, wb):
    
    # consolidate results to one book
    book = consolidate_books(results)
    
    # rename Positions to DimPositions
    book = rename_tab(book)
    
    # write book to excel_file    
    for tab in book:
        xl_utils.add_df_to_excel(book[tab], wb, tab=tab)
    
    return book

##################################################     


# check if all data is available
def check_input(DATA):
    positions = DATA['Positions']
    positions['SecurityName'] = positions['SecurityName'].apply(lambda x: x[:50] if x is not None else x)
    positions['Ticker'] = positions['Ticker'].apply(lambda x: x[:50] if x is not None else x)
    DATA['Positions'] = positions
    
    
    
############################################################################
# from api import app
# app.app_context().push()
# dataframes=results
def process_dataframes_without_report_id(dataframes):
    new_dataframes = {}
    for key, inner_dict in dataframes.items():
        for df_key, df in inner_dict.items():
            # print(df_key)

            # Replace spaces and hyphens in dataframe name with underscores
            new_key = df_key.replace(' ', '_').replace('-', '_')
            
            # Replace NaN and inf values with None
            #df.replace([np.nan, np.inf, -np.inf], None, inplace=True)
            df = df.replace({np.nan:None})
            #df = df.replace({np.inf:None})
            #df = df.replace({-np.inf:None})
            
            # Clean up column names
            new_columns = [col.replace(' ', '_').replace('-', '_').replace('$','').replace('(','').replace(')','').replace('/','_') for col in df.columns]
            df.columns = new_columns

            if key not in new_dataframes:
                new_dataframes[key] = {}
            new_dataframes[key][new_key] = df

    return new_dataframes

def test1(results):
    for key, df in results.items():
        print(key)
        print(df.columns)

# dataframes=results
def process_dataframes_with_report_id(dataframes, report_id):
    new_dataframes = {}
    for key in dataframes.keys():
        # print(key)
        df = dataframes[key]
        df['report_id'] = report_id
        new_dataframes[key] = df

    return new_dataframes

def process_nonetype_dataframes(dataframes):
    return {k: v for k, v in dataframes.items() if v is not None}


# compare if the data generated by engine are exactly same as the last time.
def dataframes_are_equal(df1, df2):
    try:
        pd.testing.assert_frame_equal(df1, df2, check_dtype=False)
        return True
    except AssertionError:
        return False

#compare if inner dictionaries are equal
def nested_dictionaries_are_equal(nested_dict1, nested_dict2):
    if nested_dict1.keys() != nested_dict2.keys():
        return False
    for key in nested_dict1.keys():
        if not dataframes_are_equal(nested_dict1[key], nested_dict2[key]):
            return False
    return True

#compare the big dictionary are equal
def dictionaries_are_equal(dict1, dict2):
    if dict1.keys() != dict2.keys():
        return False
    for key in dict1.keys():
        if not nested_dictionaries_are_equal(dict1[key], dict2[key]):
            return False
    return True

prev_file_name = None
prev_df_dict = None
file_read_count = 0

def rename_columns_with_custom_names(dataframes, keys_to_rename, new_column_names):
    num_columns = len(new_column_names)
    new_dataframes = {}
    
    for outer_key, inner_dict in dataframes.items():
        new_inner_dict = {}
        for inner_key, df in inner_dict.items():
            # Check if the DataFrame should be renamed
            if inner_key in keys_to_rename:
                if len(df.columns) < num_columns:
                    raise ValueError(f"DataFrame {inner_key} in {outer_key} has fewer than {num_columns} columns.")
                
                # Create a dictionary to map old column names to new column names
                rename_dict = {old_name: new_name for old_name, new_name in zip(df.columns[:num_columns], new_column_names)}
                
                # Rename the columns
                df.rename(columns=rename_dict, inplace=True)
            
            new_inner_dict[inner_key] = df
        
        new_dataframes[outer_key] = new_inner_dict
    
    return new_dataframes

# report_description='test'
# report_id = port_id
def insert_results_to_db(results, report_id, username, report_description):
    
    results = process_nonetype_dataframes(results)
    
    try:
        report_id = db_utils.create_report(report_id, username, report_description)
    
        results_processed = process_dataframes_with_report_id(results, report_id)
        
        logger.info('saving power bi reports to database ...')
        insert_dataframe_dic(results_processed)
        logger.info(f'report_id: {report_id} has been saved successfully!')
    except Exception as e:
        error = f'Error when inserting pbi report to database: {str(e)}'
        logger.error(error)
        raise error

    return report_id

# name = 'dm_fact_d_Positions'
# df = results_processed[name]

def insert_dataframe_dic(results_processed):
    conn = ms_sql_server.create_connection()
    for name, df in results_processed.items():
        print(name)
        ms_sql_server.insert_df(name, df, 'report_id', conn)
    conn.close()

# delete data from all pbi tables by excuting procedure DeleteReportData()
# report_id_list = [1207]
def delete_report_from_db(report_id_list):
    try: 
        conn = ms_sql_server.create_connection()
        with conn.cursor() as cursor:
            for report_id in report_id_list:
                cursor.execute("{CALL DeleteReportData(?)}", report_id)
        conn.commit()
        conn.close()
        str_id_list = ','.join([str(x) for x in report_id_list])
        print(f"report: {str_id_list} has been deleted from database")
    except Exception as e:
        print(f"Error: {e}")
    
############################################################################################################################
# Test


def test(file_name):
    
    
    #wb = xw.Book('powerbi.test.xlsx')
    wb = xw.Book('VaRCalculator.xlsm')
    params = eng.read_params(wb)
    positions = eng.read_positions(wb)

    # run engine
    DATA = eng.calc_VaR(positions, params)
    
    # pbi data
    results = generate_report(DATA) 

    # insert results to database    
    insert_results_to_db(results, 'test', 'test portfolio')

    # delete report
    report_id_list = [1214]
    delete_report_from_db(report_id_list)


    











