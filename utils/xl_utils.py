# -*- coding: utf-8 -*-
"""
Created on Tue Jan  2 16:17:16 2024

@author: mgdin
"""
import pandas as pd
import os
from utils import df_utils

# read data from excel
def read_df_from_excel(wb, tab=None, addr='A1', index=False):
    if wb is None:
        return None

    if tab is None:
        sht = wb.sheets[0]
    else:
        sht = wb.sheets[tab]

    return sht.range(addr).options(pd.DataFrame, expand='table', index=index).value

def read_dict(wb, tab):
    df = read_df_from_excel(wb, tab, index=True)
    return df.iloc[:,0] .to_dict()

# write data to excel
def add_df_to_excel(df, wb, tab=None, index=True, addr='A1'):
    if wb is None:
        return
    
    if tab is None: # write to 'Sheet1'
        tab = wb.sheets[0].name
    else:
        if wb.sheets[0].name == 'Sheet1': 
            wb.sheets[0].name = tab # retname 'Sheet1'
            
    if not tab in [sht.name for sht in wb.sheets]:
        wb.sheets.add(tab, after=wb.sheets[-1])
        
    if index:
        if df.index.name is None:
           df.index.name = 'index'
        
    sht = wb.sheets[tab]
    sht.range(addr).expand('table').clear_contents()
    sht.range(addr).options(index=index).value = df

def add_dict_to_excel(params, wb, tab='Parameters', addr='A1'):
    df = df_utils.dict_to_df(params)
    add_df_to_excel(df, wb, tab=tab, index=True, addr=addr)
    
# read a csv file and add to wb
def add_csv_to_excel(csv_file, wb, tab):
    df = pd.read_csv(csv_file)
    add_df_to_excel(df, wb, tab)

##########################################################
# pandas read
# pip install pandas openpyxl
from openpyxl import load_workbook
from zipfile import ZipFile
from xml.etree import ElementTree as ET

def pd_read_excel():

    # Load the Excel file
    df = pd.read_excel('path_to_your_file.xlsx')
    
    # Load a specific sheet by name
    df = pd.read_excel('path_to_your_file.xlsx', sheet_name='Sheet1')
    
    # Load a specific sheet by index (0 for the first sheet, 1 for the second, etc.)
    df = pd.read_excel('path_to_your_file.xlsx', sheet_name=0)

    # Load all sheets
    dfs = pd.read_excel('path_to_your_file.xlsx', sheet_name=None)
    
    # Access a specific dataframe by sheet name
    df_sheet1 = dfs['Sheet1']

    # first two rows to be the header, first column to be the index
    df = pd.read_excel('path_to_your_file.xlsx', sheet_name='Prices', header=[0,1], index_col=0)
    
    # specify data type
    df = pd.read_excel('path_to_your_file.xlsx', sheet_name='sheet', parse_dates=['TradeDate'], 
                       dtype={'ISIN': str, 'CUSIP': str, 'expected_return': 'float64', 'payment_frequency': 'int64'} )
    
# write df to excel    
def pd_to_excel():
    excel_file = r'test_data/Model-1.xlsx'
    sheet_name = 'TestSheet'       # Specify the sheet name

    df = pd.DataFrame({'name': ['John', 'Mike', 'Joe', 'Steve'], 'age': [1,2,3,4]})
    df.to_excel(excel_file, sheet_name=sheet_name, index=False)
    

def has_sheet(excel_file, sheet_name):
    try:
        excel_file.seek(0)
    except Exception:
        pass

    with ZipFile(excel_file) as z, z.open('xl/workbook.xml') as f:
        tree = ET.parse(f)
        ns = {'n': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        names = [s.get('name') for s in tree.findall('.//n:sheets/n:sheet', ns)]
        return sheet_name in names
        
###############################################################
# A simple version of write book to excel_file    

def write_book_to_xl2(book, excel_file):
    with pd.ExcelWriter(excel_file) as writer:
        for tab in book:
            book[tab].to_excel(writer, sheet_name=tab, index=False)


# write book to excel_file    
# book is a dict of name: df pair
def write_book_to_xl(book, excel_file):
    
    # if sheet_name exists, delete it
    delete_sheets(excel_file, sheet_names=book.keys())

    # Check if the file exists
    if not os.path.exists(excel_file):
        # Create an empty DataFrame and save it to initialize the file
        pd.DataFrame().to_excel(excel_file)

    # write book to excel_file    
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        for sheet_name in book:
            book[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            
    
def read_book_xl(excel_file):

    workbook = load_workbook(excel_file, data_only=True)
    book = {}
    for sheet_name in workbook.sheetnames:    
        if sheet_name in ['Sheet1']:
            continue
        # print(f'reading sheet: {sheet_name}')
        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
        book[sheet_name] = df
    return book
        
        
def delete_sheets(excel_file, sheet_names):
    # if sheet_name exists, delete it
    try:
        workbook = load_workbook(excel_file)
        # If the sheet already exists, remove it
    
        for sheet_name in sheet_names:        
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
        # Save the workbook after deleting the sheet to avoid conflicts
        workbook.save(excel_file)
    except FileNotFoundError:
        # If the file does not exist, it will be created later when writing
        pass        


    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    