# -*- coding: utf-8 -*-
"""
dashboard/upload_portfolio.py — File saving, DB insert, and processing for portfolio upload.

Public API:
    get_portfolio_file_path(client_id, filename) -> Path
    upload_portfolio(username, name, request, account_id) -> dict
    process_portfolio(file_path, port_id) -> None  (see process_uploaded_portfolio.py)
    generate_input_template(account_id, client_id=None) -> (positions_df, params, limit, client_id)
    save_portfolio_to_template(positions, params, limit, client_id, filename) -> Path
    clone_portfolio(port_id, new_port_name, username, target_weights=None, background=True) -> int
"""
from __future__ import annotations

import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import datetime
import logging
import threading
from pathlib import Path

import openpyxl
import pandas as pd
from werkzeug.utils import secure_filename

from database2 import pg_connection
from trg_config import config
from dashboard.db_port_position_var import insert_port_position_var
from dashboard.process_uploaded_portfolio import process_portfolio

logger = logging.getLogger(__name__)


def get_portfolio_file_path(client_id: int, filename: str) -> Path:
    """Return the canonical path for a portfolio file: CLIENT_DIR/<client_id>/<filename>."""
    return config['CLIENT_DIR'] / str(client_id) / filename


def _versioned_path(path: Path) -> Path:
    """If path exists, return path with .v1, .v2, ... inserted before the extension."""
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    version = 1
    while True:
        candidate = path.parent / f'{stem}.v{version}{suffix}'
        if not candidate.exists():
            return candidate
        version += 1


def _get_client_id(username: str) -> int:
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT client_id FROM "user" WHERE username = %s', (username,))
            row = cur.fetchone()
    if not row:
        raise Exception(f'User not found: {username}')
    return row[0]


def _save_file(file, client_id: int) -> tuple[Path, str]:
    """Save file to CLIENT_DIR/<client_id>/. Appends _v1, _v2, ... if filename already exists."""
    filename = secure_filename(file.filename)
    if not filename:
        raise Exception('Invalid filename')

    stem, ext = filename.rsplit('.', 1) if '.' in filename else (filename, '')
    dest = get_portfolio_file_path(client_id, filename)
    dest.parent.mkdir(parents=True, exist_ok=True)
    version = 1
    while dest.exists():
        versioned = f'{stem}_v{version}.{ext}' if ext else f'{stem}_v{version}'
        dest = dest.parent / versioned
        version += 1

    file.save(dest)
    logger.info(f'saved portfolio file: {dest}')
    return dest, dest.name


def _insert_portfolio(username: str, name: str, filename: str,
                      port_type: str | None = None,
                      description: str | None = None,
                      account_id: int | None = None) -> dict:
    """Insert a portfolio_info record and return the new portfolio entry."""
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT user_id, client_id, firstname, lastname FROM "user" WHERE username = %s', (username,))
            row = cur.fetchone()
            user_id    = row[0] if row else None
            client_id  = row[1] if row else None
            first      = (row[2] or '').strip() if row else ''
            last       = (row[3] or '').strip() if row else ''
            created_by = f'{first} {last}'.strip() or username

            cur.execute(
                """
                INSERT INTO portfolio_info
                    (port_name, filename, status, created_by, created_user_id,
                     create_date, update_date, account_id, upload_dt, client_id, port_group_id,
                     port_type, description)
                VALUES (%s, %s, 'Pending', %s, %s, NOW()::date, NOW()::date, %s, NOW(), %s, NULL,
                        %s, %s)
                RETURNING port_id, upload_dt
                """,
                (name, filename, created_by, user_id, account_id, client_id, port_type, description),
            )
            port_id, upload_dt = cur.fetchone()
        conn.commit()
    port = {
        'id':          str(port_id),
        'name':        name,
        'file':        filename,
        'upload_dt':   upload_dt.strftime('%Y-%m-%d %H:%M') if upload_dt else '—',
        'as_of_date':  '—',
        'uploaded_by': created_by,
        'mv':          None,
        'positions':   None,
        'status':      'Pending',
    }

    return port_id, port

def _update_portfolio_status(port_id: int, status: str, message: str | None = None) -> None:
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'UPDATE portfolio_info SET status = %s, message = %s WHERE port_id = %s',
                (status, message, port_id),
            )
        conn.commit()


def _run_in_background(port_id: int, port_name: str, file_path: Path) -> None:
    from api import app
    _update_portfolio_status(port_id, 'Processing')
    try:
        with app.app_context():
            process_portfolio(file_path, port_id)
        _update_portfolio_status(port_id, 'Success')
    except Exception as e:
        logger.error(f'process_portfolio failed for port_id={port_id}: {e}')
        _update_portfolio_status(port_id, 'Error', str(e))

#
# called from /api/portfolios/upload in routes.py
def upload_portfolio(username: str, name: str, request, account_id: int,
                     port_type: str | None = None,
                     description: str | None = None) -> dict:
    """Save the uploaded file and insert a portfolio_info record.

    Looks up client_id from the user table, saves the file to
    CLIENT_DIR/<client_id>/, then inserts into portfolio_info.
    """
    if 'file' not in request.files:
        raise Exception('No file part')
    file = request.files['file']
    if file.filename == '':
        raise Exception('No selected file')

    client_id = _get_client_id(username)
    file_path, filename = _save_file(file, client_id)
    insert_account_id = account_id if port_type == 'tracked' else None
    port_id, port = _insert_portfolio(username, name, filename, port_type, description, insert_account_id)

    threading.Thread(
        target=_run_in_background,
        args=(port_id, name, file_path),
        daemon=True,
    ).start()

    return port


# process_portfolio is imported from dashboard.process_uploaded_portfolio


# ── Generate / save portfolio template ───────────────────────────────────────

_POSITIONS_SHEET  = '1. Positions'
_PARAMETERS_SHEET = '3. Parameters'
_LIMIT_SHEET      = '4. Limit'

_WEIGHT_KEY_TO_CLASS = {
    'fi':  'Fixed Income',
    'eq':  'Equity',
    'alt': 'Alternatives',
    'ma':  'Multi-Asset',
    'mm':  'Money Market',
}

_DB_TO_TEMPLATE_COLS = {
    'pos_id':       'ID',
    'security_name':'SecurityName',
    'isin':         'ISIN',
    'cusip':        'Cusip',
    'ticker':       'Ticker',
    'quantity':     'Quantity',
    'market_value': 'Market Value',
    'asset_class':  'Asset Class',
    'currency':     'Currency',
}


def generate_input_template(account_id: int, client_id: int | None = None) -> tuple:
    """Fetch portfolio data for account_id from DB and return in template format.

    Returns (positions_df, params_dict, limit_dict, client_id).
    positions_df columns match the '1. Positions' sheet header.
    params_dict keys match '3. Parameters' col A labels (spaces stripped).
    limit_dict keys match '4. Limit' col A labels.
    """
    # ── 1. Resolve client_id ──────────────────────────────────────────────────
    if client_id is None:
        with pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT client_id FROM account WHERE account_id = %s', (account_id,))
                row = cur.fetchone()
        if not row:
            raise Exception(f'account not found: account_id={account_id}')
        client_id = row[0]

    # ── 2. Fetch latest positions ─────────────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT pos_id, security_name, isin, cusip, ticker,
                       quantity, market_value, "class" AS asset_class, currency, as_of_date
                FROM position_var
                WHERE account_id = %s
                  AND as_of_date = (SELECT MAX(as_of_date) FROM position_var WHERE account_id = %s)
                ORDER BY pos_id
                """,
                (account_id, account_id),
            )
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]

    df = pd.DataFrame(rows, columns=cols)
    as_of_date = df['as_of_date'].iloc[0] if not df.empty else None
    positions = df.drop(columns=['as_of_date']).rename(columns=_DB_TO_TEMPLATE_COLS)
    for col in ('Quantity', 'Market Value'):
        if col in positions.columns:
            positions[col] = pd.to_numeric(positions[col], errors='coerce')

    # ── 3. Fetch account parameters ───────────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT risk_horizon, risk_measure, base_currency, benchmark, exp_return '
                'FROM account_parameters WHERE account_id = %s ORDER BY updated_at DESC LIMIT 1',
                (account_id,),
            )
            param_row = cur.fetchone()

    db_params = dict(zip(
        ['risk_horizon', 'risk_measure', 'base_currency', 'benchmark', 'exp_return'],
        param_row,
    )) if param_row else {}
    params = {
        'AsofDate':        as_of_date,
        'ReportDate':      as_of_date,
        'RiskHorizon':     db_params.get('risk_horizon'),
        'TailMeasure':     db_params.get('risk_measure'),
        'ReturnFrequency': 'Daily',
        'Benchmark':       db_params.get('benchmark'),
        'ExpectedReturn':  db_params.get('exp_return'),
        'BaseCurrency':    db_params.get('base_currency'),
    }

    # ── 4. Fetch limits (label → value) ──────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT lc.category_label, al.limit_value
                FROM account_limit al
                JOIN limit_category lc USING (limit_category)
                WHERE al.account_id = %s
                """,
                (account_id,),
            )
            limit = {label: value for label, value in cur.fetchall()}

    return positions, params, limit, client_id


def _apply_scaler(positions: pd.DataFrame, scaler: dict) -> pd.DataFrame:
    """Scale Quantity and Market Value by asset class. Unmatched classes are unchanged."""
    positions = positions.copy()
    for asset_class, factor in scaler.items():
        mask = positions['Asset Class'] == asset_class
        positions.loc[mask, 'Quantity']     = positions.loc[mask, 'Quantity']     * factor
        positions.loc[mask, 'Market Value'] = positions.loc[mask, 'Market Value'] * factor
    return positions


def _compute_scaler(positions: pd.DataFrame, target_weights: dict) -> dict:
    """Compute per-asset-class multiplier from target percentages vs current MV breakdown.

    target_weights: {key: pct} e.g. {'fi': 40, 'eq': 35, ...} — keys from _WEIGHT_KEY_TO_CLASS.
    Returns {class_name: factor} ready for _apply_scaler().
    """
    mv = pd.to_numeric(positions['Market Value'], errors='coerce')
    total_mv = mv.sum()
    if total_mv == 0:
        return {}
    scaler = {}
    for key, target_pct in target_weights.items():
        class_name = _WEIGHT_KEY_TO_CLASS.get(key)
        if not class_name:
            continue
        mask = positions['Asset Class'] == class_name
        current_mv = pd.to_numeric(positions.loc[mask, 'Market Value'], errors='coerce').sum()
        current_pct = current_mv / total_mv * 100
        if current_pct > 0:
            scaler[class_name] = target_pct / current_pct
    return scaler


def save_portfolio_to_template(positions: pd.DataFrame, params: dict, limit: dict,
                                client_id: int, filename: str) -> Path:
    """Fill the input template with portfolio data and save to CLIENT_DIR/<client_id>/filename.

    positions columns must match the '1. Positions' sheet header.
    params keys must match '3. Parameters' col A labels (spaces stripped).
    limit keys must match '4. Limit' col A labels.
    """
    template_path = config['PUBLIC_DIR'] / 'input_template.xlsx'
    if not template_path.exists():
        raise Exception(f'template not found: {template_path}')
    wb = openpyxl.load_workbook(template_path)

    # ── Fill positions sheet ──────────────────────────────────────────────────
    ws_pos = wb[_POSITIONS_SHEET]
    header = {cell.value: cell.column for cell in ws_pos[1]}
    for row in ws_pos.iter_rows(min_row=2, max_row=ws_pos.max_row):
        for cell in row:
            cell.value = None
    for r_idx, row_dict in enumerate(positions.to_dict('records'), start=2):
        for col_name, col_idx in header.items():
            if col_name in row_dict:
                ws_pos.cell(row=r_idx, column=col_idx).value = row_dict[col_name]

    # ── Fill parameters sheet ─────────────────────────────────────────────────
    ws_par = wb[_PARAMETERS_SHEET]
    for row in ws_par.iter_rows(min_row=2, max_row=ws_par.max_row, max_col=2):
        key_cell, val_cell = row[0], row[1]
        if key_cell.value is None:
            continue
        key = str(key_cell.value).replace(' ', '')
        val = params.get(key)
        if val is not None:
            val_cell.value = val

    # ── Fill limit sheet ──────────────────────────────────────────────────────
    ws_lim = wb[_LIMIT_SHEET]
    for row in ws_lim.iter_rows(min_row=2, max_row=ws_lim.max_row, max_col=2):
        label_cell, val_cell = row[0], row[1]
        if label_cell.value is None:
            continue
        label = str(label_cell.value).strip()
        if label in limit:
            val_cell.value = limit[label]

    out_path = _versioned_path(get_portfolio_file_path(client_id, filename))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info(f'saved portfolio template: {out_path}')
    return out_path


# ── Clone portfolio ───────────────────────────────────────────────────────────

# DB column name → VaR engine column name (input-side only; VaR outputs are recomputed)
_DB_TO_ENGINE = {
    'security_id':            'SecurityID',
    'security_name':          'SecurityName',
    'isin':                   'ISIN',
    'cusip':                  'CUSIP',
    'ticker':                 'Ticker',
    'quantity':               'Quantity',
    'market_value':           'MarketValue',
    'currency':               'Currency',
    'last_price':             'LastPrice',
    'last_price_date':        'LastPriceDate',
    'asset_class':            'AssetClass',
    'asset_type':             'AssetType',
    'class':                  'Class',
    'sc1':                    'SC1',
    'sc2':                    'SC2',
    'country':                'Country',
    'region':                 'Region',
    'sector':                 'Sector',
    'industry':               'Industry',
    'expected_return':        'ExpectedReturn',
    'coupon_rate':            'CouponRate',
    'option_type':            'OptionType',
    'option_strike':          'OptionStrike',
    'payment_frequency':      'PaymentFrequency',
    'maturity_date':          'MaturityDate',
    'underlying_security_id': 'UnderlyingSecurityID',
    'underlying_id':          'UnderlyingID',
    'underlying_price':       'UnderlyingPrice',
    'risk_free_rate':         'RiskFreeRate',
    'tenor':                  'Tenor',
    'delta':                  'DELTA',
    'gamma':                  'GAMMA',
    'vega':                   'VEGA',
    'iv':                     'IV',
    'ir_tenor':               'IR_Tenor',
    'yield':                  'Yield',
    'duration':               'Duration',
    'convexity':              'Convexity',
    'ir_pv01':                'IR_PV01',
    'sp_pv01':                'SP_PV01',
    'spread_duration':        'SpreadDuration',
    'spread_convexity':       'SpreadConvexity',
    'delta_var':              'DELTA VaR',
    'ir_var':                 'IR VaR',
    'spread_var':             'SPREAD VaR',
    'gamma_var':              'GAMMA VaR',
    'total_cost':             'total_cost',
    'is_option':              'is_option',
    'broker_name':            'BrokerName',
    'broker_account':         'BrokerAccount',
}

# Non-VaR columns present in both port_position_var and position_var
_CLONE_POSITION_COLS = """
    pos_id, as_of_date,
    security_id, security_name, isin, cusip, ticker, broker_name, broker_account,
    quantity, market_value, total_cost, currency, last_price, last_price_date,
    asset_class, asset_type, "class", sc1, sc2,
    country, region, sector, industry,
    expected_return, coupon_rate, option_type, option_strike,
    payment_frequency, maturity_date, underlying_security_id, underlying_id,
    underlying_price, is_option, excluded, exclude_reason,
    risk_free_rate, tenor, delta, gamma, vega, iv,
    ir_tenor, "yield", duration, convexity,
    ir_pv01, sp_pv01, spread_duration, spread_convexity,
    delta_var, ir_var, spread_var, gamma_var
"""


_CLONE_NUMERIC_DB_COLS = frozenset({
    'quantity', 'market_value', 'total_cost', 'last_price',
    'expected_return', 'coupon_rate', 'option_strike', 'underlying_price',
    'risk_free_rate', 'tenor', 'delta', 'gamma', 'vega', 'iv',
    'ir_tenor', 'yield', 'duration', 'convexity',
    'ir_pv01', 'sp_pv01', 'spread_duration', 'spread_convexity',
    'delta_var', 'ir_var', 'spread_var', 'gamma_var',
})


def _fetch_clone_positions(sql: str, params: tuple) -> tuple[pd.DataFrame, object]:
    """Execute sql, rename DB columns to engine convention, and return (df, asof_date)."""
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
    if not rows:
        return pd.DataFrame(), None
    df   = pd.DataFrame(rows, columns=cols)
    asof = df['as_of_date'].iloc[0]
    # Convert psycopg2 Decimal → float before rename so arithmetic works downstream
    for col in _CLONE_NUMERIC_DB_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    df   = df.drop(columns=['as_of_date']).rename(columns=_DB_TO_ENGINE)
    df['SecurityID'] = df['SecurityID'].apply(
        lambda x: None if (x is None or (isinstance(x, float) and pd.isna(x)))
                  else (str(int(x)) if isinstance(x, float) else str(x))
    )
    return df, asof


def _build_positions_from_adhoc(src_port_id: int) -> tuple:
    """Read enriched positions + params from port_position_var / port_parameters."""
    sql = f'SELECT {_CLONE_POSITION_COLS} FROM port_position_var WHERE port_id = %s ORDER BY pos_id'
    positions, asof_date = _fetch_clone_positions(sql, (src_port_id,))
    if positions.empty:
        raise Exception(f'no positions in port_position_var for port_id={src_port_id}')

    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT "PortfolioName", "AsofDate", "ReportDate", "RiskHorizon", "TailMeasure", '
                '"ReturnFrequency", "Benchmark", "ExpectedReturn", "BaseCurrency" '
                'FROM port_parameters WHERE port_id = %s',
                (src_port_id,),
            )
            row = cur.fetchone()

    if row:
        params = dict(zip(
            ['PortfolioName', 'AsofDate', 'ReportDate', 'RiskHorizon', 'TailMeasure',
             'ReturnFrequency', 'Benchmark', 'ExpectedReturn', 'BaseCurrency'],
            row,
        ))
    else:
        params = {'AsofDate': asof_date, 'ReportDate': asof_date}

    return positions, params, asof_date


def _build_positions_from_tracked(account_id: int) -> tuple:
    """Read latest enriched positions + params from position_var / account_parameters."""
    sql = f"""
        SELECT {_CLONE_POSITION_COLS}
        FROM position_var
        WHERE account_id = %s
          AND as_of_date = (SELECT MAX(as_of_date) FROM position_var WHERE account_id = %s)
        ORDER BY pos_id
    """
    positions, asof_date = _fetch_clone_positions(sql, (account_id, account_id))
    if positions.empty:
        raise Exception(f'no positions in position_var for account_id={account_id}')

    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT risk_horizon, risk_measure, base_currency, benchmark, exp_return '
                'FROM account_parameters WHERE account_id = %s ORDER BY updated_at DESC LIMIT 1',
                (account_id,),
            )
            row = cur.fetchone()

    db_p = dict(zip(
        ['risk_horizon', 'risk_measure', 'base_currency', 'benchmark', 'exp_return'],
        row,
    )) if row else {}
    params = {
        'AsofDate':        asof_date,
        'ReportDate':      asof_date,
        'RiskHorizon':     db_p.get('risk_horizon'),
        'TailMeasure':     db_p.get('risk_measure'),
        'ReturnFrequency': 'Daily',
        'Benchmark':       db_p.get('benchmark'),
        'ExpectedReturn':  db_p.get('exp_return'),
        'BaseCurrency':    db_p.get('base_currency'),
    }

    return positions, params, asof_date


def _run_var_on_positions(positions: pd.DataFrame, port_id: int, asof_date, port_name: str) -> None:
    """Run VaR pipeline steps 3-7 on pre-enriched positions and write to port_position_var."""
    from process2 import var_engine
    from process2.calculate_var import add_beta_to_result, fetch_betas_bulk

    if 'excluded' in positions.columns:
        active   = positions[positions['excluded'] != True].reset_index(drop=True)
        excluded = positions[positions['excluded'] == True].reset_index(drop=True)
    else:
        active   = positions.copy()
        excluded = pd.DataFrame(columns=positions.columns)
    logger.info(f'[{port_name}] VaR split: {len(active)} active, {len(excluded)} excluded')

    var_metrics = var_engine.calc_var(active)
    result      = active.set_index('pos_id').join(var_metrics, how='left').reset_index()

    if not excluded.empty:
        excl = excluded.copy()
        for col in var_metrics.columns:
            excl[col] = None
        result = pd.concat([result, excl], ignore_index=True)

    sec_ids    = result['SecurityID'].dropna().unique().tolist()
    betas_bulk = fetch_betas_bulk(['SP500_1Y'], sec_ids)
    betas      = betas_bulk.get('SP500_1Y', {})
    result     = add_beta_to_result(result, betas, logger)

    n = insert_port_position_var(result, port_id, asof_date)
    logger.info(f'[{port_name}] inserted {n} rows into port_position_var')


def _clone_in_background(port_id: int, port_name: str,
                          positions: pd.DataFrame, params: dict, asof_date) -> None:
    from api import app
    from dashboard.db_port_input import insert_port_parameters, insert_port_positions

    _update_portfolio_status(port_id, 'Processing')
    try:
        with app.app_context():
            insert_port_parameters({**params, 'PortfolioName': port_name}, port_id)
            n_pos = insert_port_positions(positions, port_id)
            logger.info(f'[{port_name}] inserted params and {n_pos} positions for port_id={port_id}')

            mv = float(pd.to_numeric(positions['MarketValue'], errors='coerce').sum()) \
                 if 'MarketValue' in positions.columns else None
            with pg_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        'UPDATE portfolio_info SET market_value = %s, as_of_date = %s WHERE port_id = %s',
                        (mv, asof_date, port_id),
                    )
                conn.commit()

            _run_var_on_positions(positions, port_id, asof_date, port_name)
        _update_portfolio_status(port_id, 'Success')
    except Exception as e:
        logger.error(f'clone_portfolio failed for port_id={port_id}: {e}')
        _update_portfolio_status(port_id, 'Error', str(e))


def clone_portfolio(port_id: int, new_port_name: str, username: str,
                    target_weights: dict | None = None, background: bool = True) -> int:
    """Clone a portfolio under a new name, reading directly from the database.

    For adhoc sources (account_id is None): reads from port_position_var + port_parameters.
    For tracked sources: reads from position_var + account_parameters.
    target_weights: {key: pct} e.g. {'fi': 40, 'eq': 35} — scales Quantity/MarketValue so
    the cloned portfolio matches the requested allocation percentages.
    Returns the new port_id.
    """
    # Step 1: Fetch source portfolio_info
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT account_id, client_id FROM portfolio_info WHERE port_id = %s',
                (port_id,),
            )
            row = cur.fetchone()
    if not row:
        raise Exception(f'portfolio not found: port_id={port_id}')
    src_account_id, _src_client_id = row

    # Step 2: Resolve caller's client_id and new filename
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT client_id FROM "user" WHERE username = %s', (username,))
            row = cur.fetchone()
    if not row:
        raise Exception(f'user not found: {username}')
    client_id    = row[0]
    new_filename = f'{new_port_name}.xlsx'

    # Step 3: Build positions + params from DB
    if src_account_id is None:
        positions, params, asof_date = _build_positions_from_adhoc(port_id)
    else:
        positions, params, asof_date = _build_positions_from_tracked(src_account_id)
    positions = positions.copy()

    # Step 4: Apply target weights (scale Quantity and MarketValue by AssetClass)
    if target_weights and 'AssetClass' in positions.columns and 'MarketValue' in positions.columns:
        mv_total = pd.to_numeric(positions['MarketValue'], errors='coerce').sum()
        if mv_total > 0:
            for key, target_pct in target_weights.items():
                class_name = _WEIGHT_KEY_TO_CLASS.get(key)
                if not class_name:
                    continue
                mask       = positions['AssetClass'] == class_name
                current_mv = pd.to_numeric(positions.loc[mask, 'MarketValue'], errors='coerce').sum()
                if current_mv > 0:
                    factor = (target_pct / 100 * mv_total) / current_mv
                    positions.loc[mask, 'Quantity']    = positions.loc[mask, 'Quantity']    * factor
                    positions.loc[mask, 'MarketValue'] = positions.loc[mask, 'MarketValue'] * factor

    # Step 5: Generate Excel file (write-only, for download)
    template_positions = positions.rename(columns={
        'pos_id': 'ID', 'CUSIP': 'Cusip',
        'MarketValue': 'Market Value', 'AssetClass': 'Asset Class',
    })
    new_file_path = save_portfolio_to_template(template_positions, params, {}, client_id, new_filename)

    # Step 6: Insert portfolio_info
    new_port_id, _ = _insert_portfolio(username, new_port_name, new_file_path.name, port_type='adhoc')
    logger.info(f'cloned port_id={port_id} -> new port_id={new_port_id} ({new_port_name})')

    # Step 7: Insert DB records and run VaR pipeline
    if background:
        threading.Thread(
            target=_clone_in_background,
            args=(new_port_id, new_port_name, positions, params, asof_date),
            daemon=True,
        ).start()
    else:
        _clone_in_background(new_port_id, new_port_name, positions, params, asof_date)

    return new_port_id


def clone_portfolio_old(port_id: int, new_port_name: str, username: str,
                    target_weights: dict | None = None, background: bool = True) -> int:
    """Clone a portfolio under a new name, optionally targeting new asset class weights.

    If the source has no account_id, reads portfolio data from the source file directly.
    If it has an account_id, generates fresh data from the DB via generate_input_template().
    target_weights: {key: pct} e.g. {'fi': 40, 'eq': 35, ...} — computes per-class scalers
    and applies them so the cloned portfolio matches the requested allocation percentages.
    Saves to a new file, inserts a portfolio_info row, kicks off processing.
    Returns the new port_id.
    """
    # ── 1. Fetch source portfolio ─────────────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT account_id, filename, client_id FROM portfolio_info WHERE port_id = %s',
                (port_id,),
            )
            row = cur.fetchone()
    if not row:
        raise Exception(f'portfolio not found: port_id={port_id}')
    src_account_id, src_filename, src_client_id = row

    # ── 2. Resolve caller's client_id ────────────────────────────────────────
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT client_id FROM "user" WHERE username = %s', (username,))
            row = cur.fetchone()
    if not row:
        raise Exception(f'user not found: {username}')
    client_id = row[0]

    new_filename = f'{new_port_name}.xlsx'

    # ── 3. Build portfolio data ───────────────────────────────────────────────
    if src_account_id is None:
        # No account — read raw from source file, no validation or column renaming
        src_path = get_portfolio_file_path(src_client_id, src_filename)
        positions = pd.read_excel(src_path, sheet_name=_POSITIONS_SHEET)
        par_df = pd.read_excel(src_path, sheet_name=_PARAMETERS_SHEET, usecols=[0, 1])
        params = {
            str(r.iloc[0]).replace(' ', ''): r.iloc[1]
            for _, r in par_df.iterrows() if pd.notna(r.iloc[0])
        }
        lim_df = pd.read_excel(src_path, sheet_name=_LIMIT_SHEET, usecols=[0, 1])
        limit = {
            str(r.iloc[0]).strip(): r.iloc[1]
            for _, r in lim_df.iterrows() if pd.notna(r.iloc[0])
        }
    else:
        # Has account — generate fresh data from DB
        positions, params, limit, _ = generate_input_template(src_account_id, client_id)

    # ── Apply target weights and save (unified for both cases) ───────────────
    if target_weights:
        scaler = _compute_scaler(positions, target_weights)
        if scaler:
            positions = _apply_scaler(positions, scaler)
    new_file_path = save_portfolio_to_template(positions, params, limit, client_id, new_filename)

    # ── 4. Insert portfolio_info row ──────────────────────────────────────────
    new_port_id, _ = _insert_portfolio(username, new_port_name, new_file_path.name, port_type='adhoc')
    logger.info(f'cloned port_id={port_id} -> new port_id={new_port_id} ({new_port_name})')

    # ── 5. Kick off processing ────────────────────────────────────────────────
    if background:
        threading.Thread(
            target=_run_in_background,
            args=(new_port_id, new_port_name, new_file_path),
            daemon=True,
        ).start()
    else:
        _run_in_background(new_port_id, new_port_name, new_file_path)

    return new_port_id


# ── Test ──────────────────────────────────────────────────────────────────────

def test_clone(port_id: int = 5359,
               new_port_name: str = 'TEST_CLONE_5359',
               username: str = 'test1@trg.com') -> None:

    target_weights = {
        'fi': 40,
        'eq': 35,
        'alt': 15,
        'ma': 5,
        'mm': 5,
    }
    clone_portfolio(port_id, new_port_name, username,
                    target_weights, background = False)

if __name__ == '__main__':
    import argparse
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)-8s %(message)s')
    parser = argparse.ArgumentParser(description='Test clone_portfolio step by step.')
    parser.add_argument('--port-id',       type=int, default=5359,               metavar='PORT_ID')
    parser.add_argument('--name',          type=str, default='TEST_CLONE_5359',  metavar='NAME')
    parser.add_argument('--username',      type=str, default='test1@trg.com',    metavar='USER')
    args = parser.parse_args()
    test_clone(port_id=args.port_id, new_port_name=args.name, username=args.username)
